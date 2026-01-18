# Standard library imports
from django.db.models import F
import io
import json
import os
import re
import tempfile
import time
from collections import Counter, defaultdict
from datetime import datetime, timedelta

# Django imports
from django.contrib.auth import get_user_model
from django.core.cache import cache
from django.db.models import Q, Avg, Count, Sum
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone

from django.conf import settings
from reportlab.platypus import Image, Spacer


# Django REST Framework imports
from rest_framework import status
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

# Third-party imports
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="34495E", fill_type="solid")
center_alignment = Alignment(horizontal="center", vertical="center")

subheader_fill = PatternFill(start_color="9B59B6", fill_type="solid")
left_alignment = Alignment(horizontal="left", vertical="center")


# ReportLab imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import (
    Pembayaran, Anggota, LaporanSampah, Jadwal, 
    TimPengangkut, DetailAnggotaJadwal 
)

# Import serializers
from .report import (
    KeuanganReportSerializer, KeuanganSummarySerializer, KeuanganTableSerializer,
    AnggotaReportSerializer, AnggotaSummarySerializer, AnggotaTableSerializer,
    LaporanSampahReportSerializer, LaporanSampahSummarySerializer, LaporanSampahTableSerializer,
    JadwalReportSerializer, JadwalSummarySerializer, JadwalPerTimSerializer, JadwalAnggotaSerializer,
    UserStatReportSerializer, UserStatSummarySerializer, UserStatTableSerializer,
    MonthlyReportSerializer, ExportRequestSerializer
)

# Import models
from django.contrib.auth.models import User
from apk.models import (
    Pembayaran, Anggota, LaporanSampah, Jadwal, 
    TimPengangkut, DetailAnggotaJadwal
)

class ReportViewSet(APIView):
    """Base class untuk semua reports"""
    permission_classes = [IsAuthenticated]

    def check_admin_permission(self, request):
        """Check if user is admin"""
        if not hasattr(request.user, 'role') or request.user.role != 'admin':
            raise PermissionDenied("Hanya admin yang bisa mengakses laporan")
        return True

    def parse_date_range(self, request):
        """
        Ambil start_date & end_date dari request
        Support:
        - POST body (JSON)
        - Query param (?start_date=YYYY-MM-DD)
        """
        start_date = request.data.get("start_date") or request.query_params.get("start_date")
        end_date = request.data.get("end_date") or request.query_params.get("end_date")

        if not start_date or not end_date:
            raise ValidationError("start_date dan end_date wajib diisi")

        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise ValidationError("Format tanggal harus YYYY-MM-DD")

        if start_date > end_date:
            raise ValidationError("start_date tidak boleh lebih besar dari end_date")

        return start_date, end_date

    def format_period(self, start_date, end_date):
        """Format period string"""
        return f"{start_date.strftime('%d/%m/%Y')} s/d {end_date.strftime('%d/%m/%Y')}"

class KeuanganReportView(ReportViewSet):
    """View laporan pembayaran"""

    def post(self, request):
        try:
            self.check_admin_permission(request)
            start_date, end_date = self.parse_date_range(request)

            pembayaran_qs = Pembayaran.objects.filter(
                tanggalBayar__range=[start_date, end_date]
            )

            # ===== INFO / SUMMARY =====
            summary_data = {
                "period": self.format_period(start_date, end_date),
                "total_pendapatan": pembayaran_qs.filter(
                    statusBayar="lunas"
                ).aggregate(total=Sum("jumlahBayar"))["total"] or 0,
                "total_lunas": pembayaran_qs.filter(statusBayar="lunas").count(),
                "total_pending": pembayaran_qs.filter(statusBayar="pending").count(),
                "total_gagal": pembayaran_qs.filter(statusBayar="gagal").count(),
                "metode_bayar_stats": {
                    row["metodeBayar"]: {
                        "count": row["count"],
                        "total": row["total"]
                    }
                    for row in pembayaran_qs.values("metodeBayar")
                        .annotate(count=Count("idPembayaran"), total=Sum("jumlahBayar"))
                },
                "tanggal_generate": timezone.now()
            }

            # Validasi summary dengan serializer
            summary_serializer = KeuanganSummarySerializer(data=summary_data)
            if not summary_serializer.is_valid():
                return Response(summary_serializer.errors, status=400)
            
            # ===== TABLE DATA =====
            table_qs = pembayaran_qs.select_related("idAnggota").values(
                "tanggalBayar",
                "idAnggota__nama",
                "jumlahBayar",
                "metodeBayar",
                "statusBayar"
            ).order_by("-tanggalBayar")

            table_data = [
                {
                    "tanggal_bayar": r["tanggalBayar"],
                    "nama_anggota": r["idAnggota__nama"],
                    "jumlah_bayar": r["jumlahBayar"],
                    "metode_bayar": r["metodeBayar"],
                    "status_bayar": r["statusBayar"],
                }
                for r in table_qs
            ]

            # Validasi table dengan serializer
            table_serializer = KeuanganTableSerializer(data=table_data, many=True)
            if not table_serializer.is_valid():
                return Response(table_serializer.errors, status=400)

            # Gabungkan response
            response_data = {
                "info": summary_serializer.validated_data,
                "table": table_serializer.validated_data
            }

            # Validasi final response
            report_serializer = KeuanganReportSerializer(data=response_data)
            if not report_serializer.is_valid():
                return Response(report_serializer.errors, status=400)

            return Response(report_serializer.validated_data)

        except Exception as e:
            return Response({'error': str(e)}, status=500)


class AnggotaReportView(ReportViewSet):
    """View laporan anggota"""

    def post(self, request):
        try:
            self.check_admin_permission(request)
            start_date, end_date = self.parse_date_range(request)

            anggota_qs = Anggota.objects.filter(
                tanggalStart__range=[start_date, end_date]
            ).select_related('user')

            # ===== SUMMARY =====
            total_anggota = anggota_qs.count()
            aktif = anggota_qs.filter(status='aktif').count()
            non_aktif = anggota_qs.filter(status='non-aktif').count()

            today = timezone.now().date()
            akan_expired = Anggota.objects.filter(
                status='aktif',
                tanggalEnd__range=[today, today + timedelta(days=30)]
            ).count()

            baru_bulan_ini = Anggota.objects.filter(
                tanggalStart__range=[start_date, end_date]
            ).count()

            jenis_sampah_stats = {
                i['jenisSampah'] or 'Tidak diketahui': i['count']
                for i in anggota_qs.values('jenisSampah')
                .annotate(count=Count('idAnggota'))
            }

            summary_data = {
                'period': self.format_period(start_date, end_date),
                'total_anggota': total_anggota,
                'aktif': aktif,
                'non_aktif': non_aktif,
                'akan_expired': akan_expired,
                'baru_bulan_ini': baru_bulan_ini,
                'jenis_sampah_stats': jenis_sampah_stats,
                'tanggal_generate': timezone.now()
            }

            # Validasi summary dengan serializer
            summary_serializer = AnggotaSummarySerializer(data=summary_data)
            if not summary_serializer.is_valid():
                return Response(summary_serializer.errors, status=400)
            
            # ===== TABLE DATA =====
            table_data = [
                {
                    'nama_anggota': a.nama,
                    'status': a.status,
                    'jenis_sampah': a.jenisSampah or '-',
                    'tanggal_start': a.tanggalStart,
                    'tanggal_end': a.tanggalEnd
                }
                for a in anggota_qs.order_by('-tanggalStart')
            ]

            # Validasi table dengan serializer
            table_serializer = AnggotaTableSerializer(data=table_data, many=True)
            if not table_serializer.is_valid():
                return Response(table_serializer.errors, status=400)

            # Gabungkan response
            response_data = {
                'info': summary_serializer.validated_data,
                'table': table_serializer.validated_data
            }

            # Validasi final response
            report_serializer = AnggotaReportSerializer(data=response_data)
            if not report_serializer.is_valid():
                return Response(report_serializer.errors, status=400)

            return Response(report_serializer.validated_data)

        except Exception as e:
            return Response({'error': str(e)}, status=500)


class LaporanSampahReportView(ReportViewSet):
    """View laporan sampah"""

    def post(self, request):
        try:
            self.check_admin_permission(request)
            start_date, end_date = self.parse_date_range(request)

            laporan_qs = LaporanSampah.objects.filter(
                tanggal_lapor__range=[start_date, end_date]
            )

            # ===== SUMMARY =====
            summary_data = {
                "period": self.format_period(start_date, end_date),
                "total_laporan": laporan_qs.count(),
                "pending": laporan_qs.filter(status="pending").count(),
                "proses": laporan_qs.filter(status="proses").count(),
                "selesai": laporan_qs.filter(status="selesai").count(),
                "avg_response_time": None,  # TODO: Hitung jika ada data waktu respons
                "tanggal_generate": timezone.now()
            }

            # Validasi summary dengan serializer
            summary_serializer = LaporanSampahSummarySerializer(data=summary_data)
            if not summary_serializer.is_valid():
                return Response(summary_serializer.errors, status=400)
            
            table_qs = laporan_qs.select_related("idUser").values(
                "tanggal_lapor",
                "alamat",
                "status",
                nama_pelapor=F("nama"),
            ).order_by("-tanggal_lapor")

            table_data = list(table_qs)

            # Validasi table dengan serializer
            table_serializer = LaporanSampahTableSerializer(data=table_data, many=True)
            if not table_serializer.is_valid():
                return Response(table_serializer.errors, status=400)

            # Gabungkan response
            response_data = {
                "info": summary_serializer.validated_data,
                "table": table_serializer.validated_data
            }

            # Validasi final response
            report_serializer = LaporanSampahReportSerializer(data=response_data)
            if not report_serializer.is_valid():
                return Response(report_serializer.errors, status=400)

            return Response(report_serializer.validated_data)

        except Exception as e:
            return Response({'error': str(e)}, status=500)


class JadwalReportView(ReportViewSet):
    """View laporan pengangkutan (group by tim angkut)"""

    def post(self, request):
        try:
            self.check_admin_permission(request)
            start_date, end_date = self.parse_date_range(request)

            jadwal_qs = Jadwal.objects.filter(
                tanggalJadwal__range=[start_date, end_date]
            )

            detail_qs = DetailAnggotaJadwal.objects.select_related(
                "idJadwal", "idAnggota", "idJadwal__idTim"
            ).filter(
                idJadwal__tanggalJadwal__range=[start_date, end_date]
            )

            # ===== SUMMARY =====
            summary_data = {
                "period": self.format_period(start_date, end_date),
                "total_jadwal": jadwal_qs.count(),
                "total_tim": jadwal_qs.values("idTim").distinct().count(),
                "total_anggota_terjadwal": detail_qs.count(),
                "status_stats": {
                    row["status_pengangkutan"]: row["count"]
                    for row in detail_qs.values("status_pengangkutan")
                        .annotate(count=Count("id"))
                },
                "tanggal_generate": timezone.now()
            }

            # Validasi summary dengan serializer
            summary_serializer = JadwalSummarySerializer(data=summary_data)
            if not summary_serializer.is_valid():
                return Response(summary_serializer.errors, status=400)

            # ===== GROUP BY TIM =====
            data_by_tim = defaultdict(lambda: {
                "total_jadwal": set(),
                "total_anggota": 0,
                "status_stats": defaultdict(int),
                "detail": []
            })

            for d in detail_qs:
                tim = d.idJadwal.idTim.namaTim

                data_by_tim[tim]["total_jadwal"].add(d.idJadwal_id)
                data_by_tim[tim]["total_anggota"] += 1
                data_by_tim[tim]["status_stats"][d.status_pengangkutan] += 1

                data_by_tim[tim]["detail"].append({
                    "tanggal_jadwal": d.idJadwal.tanggalJadwal,
                    "nama_anggota": d.idAnggota.nama,
                    "status_pengangkutan": d.status_pengangkutan
                })

            table_data = [
                {
                    "nama_tim": tim,
                    "total_jadwal": len(v["total_jadwal"]),
                    "total_anggota": v["total_anggota"],
                    "status_stats": dict(v["status_stats"]),
                    "detail": v["detail"]
                }
                for tim, v in data_by_tim.items()
            ]

            # Validasi table dengan serializer
            table_serializer = JadwalPerTimSerializer(data=table_data, many=True)
            if not table_serializer.is_valid():
                return Response(table_serializer.errors, status=400)

            # Gabungkan response
            response_data = {
                "info": summary_serializer.validated_data,
                "table": table_serializer.validated_data
            }

            # Validasi final response
            report_serializer = JadwalReportSerializer(data=response_data)
            if not report_serializer.is_valid():
                return Response(report_serializer.errors, status=400)

            return Response(report_serializer.validated_data)

        except Exception as e:
            return Response({'error': str(e)}, status=500)


class UserStatReportView(ReportViewSet):
    """View laporan statistik user (summary + table sortable)"""

    def get(self, request):
        try:
            self.check_admin_permission(request)
            User = get_user_model()

            # ===== SUMMARY =====
            total_users = User.objects.count()
            admin_count = User.objects.filter(role='admin').count()
            anggota_count = User.objects.filter(role='anggota').count()
            tamu_count = User.objects.filter(role='tamu').count()
            tim_angkut_count = User.objects.filter(role='tim_angkut').count()
            active_users = User.objects.filter(is_active=True).count()

            today = timezone.now().date()
            bulan_ini_start = today.replace(day=1)
            new_users_month = User.objects.filter(
                date_joined__date__gte=bulan_ini_start
            ).count()

            summary_data = {
                "total_users": total_users,
                "admin_count": admin_count,
                "anggota_count": anggota_count,
                "tamu_count": tamu_count,
                "tim_angkut_count": tim_angkut_count,
                "active_users": active_users,
                "new_users_month": new_users_month,
                "tanggal_generate": timezone.now()
            }

            # Validasi summary dengan serializer
            summary_serializer = UserStatSummarySerializer(data=summary_data)
            if not summary_serializer.is_valid():
                return Response(summary_serializer.errors, status=400)

            # ===== TABLE DATA =====
            ordering = request.query_params.get("ordering", "-date_joined")
            table_qs = User.objects.values(
                "username",
                "email",
                "role",
                "is_active",
                "date_joined"
            ).order_by(ordering)

            table_data = list(table_qs)

            # Validasi table dengan serializer
            table_serializer = UserStatTableSerializer(data=table_data, many=True)
            if not table_serializer.is_valid():
                return Response(table_serializer.errors, status=400)

            # Gabungkan response
            response_data = {
                "info": summary_serializer.validated_data,
                "table": table_serializer.validated_data
            }

            # Validasi final response
            report_serializer = UserStatReportSerializer(data=response_data)
            if not report_serializer.is_valid():
                return Response(report_serializer.errors, status=400)

            return Response(report_serializer.validated_data)

        except Exception as e:
            return Response({'error': str(e)}, status=500)


class MonthlyReportView(ReportViewSet):
    """View untuk laporan bulanan dengan data asli"""
  
    def post(self, request):
        try:
            self.check_admin_permission(request)
          
            # Validasi input
            month = request.data.get('month')
            year = request.data.get('year')
            
            if not month or not year:
                return Response({'error': 'month dan year diperlukan'}, status=400)
            
            try:
                month = int(month)
                year = int(year)
                
                if month < 1 or month > 12:
                    return Response({'error': 'Month harus antara 1-12'}, status=400)
                
                if year < 2000 or year > 2100:
                    return Response({'error': 'Year tidak valid'}, status=400)
                    
            except ValueError:
                return Response({'error': 'Month dan year harus angka'}, status=400)
            
            # Hitung range tanggal untuk bulan tersebut
            start_date = datetime(year, month, 1).date()
            if month == 12:
                end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
            
            # Get data keuangan untuk bulan ini
            keuangan_data = self.get_keuangan_data(start_date, end_date)
            
            # Get data anggota untuk bulan ini
            anggota_data = self.get_anggota_data(start_date, end_date)
            
            # Get data jadwal untuk bulan ini
            jadwal_data = self.get_jadwal_data(start_date, end_date)
            
            # Get data laporan sampah untuk bulan ini
            laporan_data = self.get_laporan_data(start_date, end_date)
            
            # Hitung success rate (simplified)
            total_services = jadwal_data.get('total_anggota_terjadwal', 0)
            completed_services = DetailAnggotaJadwal.objects.filter(
                idJadwal__tanggalJadwal__range=[start_date, end_date],
                status_pengangkutan='selesai'
            ).count()
            
            success_rate = (completed_services / total_services * 100) if total_services > 0 else 0
            
            # Hitung resolution rate
            total_reports = laporan_data.get('total_laporan', 0)
            completed_reports = laporan_data.get('selesai', 0)
            resolution_rate = (completed_reports / total_reports * 100) if total_reports > 0 else 0
            
            # PERBAIKAN: Gunakan get_user_model() untuk menghitung user
            User = get_user_model()
            
            # Nama bulan
            month_names = [
                'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
            ]
            
            # Prepare response data
            report_data = {
                'bulan': f"{month_names[month-1]} {year}",
                'tahun': year,
                'total_pendapatan': float(keuangan_data.get('total_pendapatan', 0)),
                'total_transaksi': keuangan_data.get('total_lunas', 0),
                'total_anggota': anggota_data.get('total_anggota', 0),
                'anggota_baru': anggota_data.get('baru_bulan_ini', 0),
                'anggota_expired': anggota_data.get('akan_expired', 0),
                'total_jadwal': jadwal_data.get('total_jadwal', 0),
                'anggota_dilayani': jadwal_data.get('total_anggota_terjadwal', 0),
                'success_rate': round(success_rate, 2),
                'total_laporan': laporan_data.get('total_laporan', 0),
                'resolution_rate': round(resolution_rate, 2),
                'summary': {
                    'pendapatan_per_anggota': round(
                        float(keuangan_data.get('total_pendapatan', 0)) / 
                        max(anggota_data.get('aktif', 1), 1), 
                        2
                    ),
                    'laporan_per_user': round(
                        laporan_data.get('total_laporan', 0) / 
                        max(User.objects.count(), 1),  # PERBAIKAN DI SINI
                        2
                    ),
                    'efficiency_rate': round(
                        (success_rate + resolution_rate) / 2, 
                        2
                    ) if total_reports > 0 else 0
                },
                'tanggal_generate': timezone.now()
            }
            
            # Validasi dengan serializer
            serializer = MonthlyReportSerializer(data=report_data)
            if not serializer.is_valid():
                return Response(serializer.errors, status=400)
            
            return Response(serializer.validated_data)
            
        except PermissionDenied as e:
            return Response({'error': str(e)}, status=403)
        except Exception as e:
            import traceback
            print(f"Error in MonthlyReportView: {e}")
            print(f"Traceback: {traceback.format_exc()}")
            return Response({'error': f'Internal server error: {str(e)}'}, status=500)
    
    def get_keuangan_data(self, start_date, end_date):
        """Get keuangan data for period"""
        pembayaran_qs = Pembayaran.objects.filter(
            tanggalBayar__range=[start_date, end_date]
        )
        
        aggregation = pembayaran_qs.filter(statusBayar='lunas').aggregate(
            total_pendapatan=Sum('jumlahBayar')
        )
        
        total_pendapatan = aggregation['total_pendapatan'] or 0
        total_lunas = pembayaran_qs.filter(statusBayar='lunas').count()
        
        return {
            'total_pendapatan': total_pendapatan,
            'total_lunas': total_lunas
        }
    
    def get_anggota_data(self, start_date, end_date):
        """Get anggota data for period"""
        anggota_qs = Anggota.objects.all()
        
        today = timezone.now().date()
        future_30 = today + timedelta(days=30)
        
        return {
            'total_anggota': anggota_qs.count(),
            'aktif': anggota_qs.filter(status='aktif').count(),
            'baru_bulan_ini': Anggota.objects.filter(
                tanggalStart__range=[start_date, end_date]
            ).count(),
            'akan_expired': Anggota.objects.filter(
                status='aktif',
                tanggalEnd__range=[today, future_30]
            ).count()
        }
    
    def get_jadwal_data(self, start_date, end_date):
        """Get jadwal data for period"""
        jadwal_qs = Jadwal.objects.filter(
            tanggalJadwal__range=[start_date, end_date]
        )
        
        total_anggota_terjadwal = DetailAnggotaJadwal.objects.filter(
            idJadwal__tanggalJadwal__range=[start_date, end_date]
        ).count()
        
        return {
            'total_jadwal': jadwal_qs.count(),
            'total_anggota_terjadwal': total_anggota_terjadwal
        }
    
    def get_laporan_data(self, start_date, end_date):
        """Get laporan data for period"""
        laporan_qs = LaporanSampah.objects.filter(
            tanggal_lapor__range=[start_date, end_date]
        )
        
        return {
            'total_laporan': laporan_qs.count(),
            'selesai': laporan_qs.filter(status='selesai').count()
        }
    
class ExportReportView(ReportViewSet):
    """View untuk export laporan ke berbagai format dengan PDF/Excel support"""
    
    def post(self, request):
        try:
            self.check_admin_permission(request)
            
            # Validasi input
            report_type = request.data.get('report_type')
            format_type = request.data.get('format', 'json')
            filters = request.data.get('filters', {})
            
            if not report_type:
                return Response({'error': 'report_type diperlukan'}, status=400)
            
            # Dispatch ke report generator yang sesuai
            report_data = self.get_report_data(report_type, request, filters)
            
            if isinstance(report_data, Response):
                return report_data
            
            # Generate response berdasarkan format
            if format_type == 'json':
                return Response(report_data)
                
            elif format_type == 'pdf':
                return self.generate_pdf_response(report_data, report_type, filters)
                
            elif format_type == 'excel':
                return self.generate_excel_response(report_data, report_type, filters)
                
            else:
                return Response({'error': 'Format tidak didukung'}, status=400)
            
        except PermissionDenied as e:
            return Response({'error': str(e)}, status=403)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({'error': f'Internal server error: {str(e)}'}, status=500)
    
    def get_report_data(self, report_type, request, filters):
        """Get report data based on type"""
        # Update request data dengan filters
        if filters:
            for key, value in filters.items():
                if key not in request.data:
                    request.data[key] = value
        
        if report_type == 'keuangan':
            view = KeuanganReportView()
            view.request = request
            response = view.post(request)
            
        elif report_type == 'anggota':
            view = AnggotaReportView()
            view.request = request
            response = view.post(request)
            
        elif report_type == 'laporan-sampah':
            view = LaporanSampahReportView()
            view.request = request
            response = view.post(request)
            
        elif report_type == 'jadwal':
            view = JadwalReportView()
            view.request = request
            response = view.post(request)
            
        elif report_type == 'user-stats':
            view = UserStatReportView()
            view.request = request
            response = view.get(request)
            
        elif report_type == 'monthly':
            view = MonthlyReportView()
            view.request = request
            response = view.post(request)

        elif report_type == 'dampak-lingkungan':
            view = DampakLingkunganReportView()
            view.request = request
            response = view.post(request)
            
        else:
            return Response({'error': 'Jenis report tidak valid'}, status=400)
        
        if hasattr(response, 'data'):
            return response.data
        return {}
    
    def generate_pdf_response(self, data, report_type, filters):
        """Generate PDF response"""
        buffer = io.BytesIO()
        
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        elements = []
        styles = getSampleStyleSheet()

        # ===== TAMBAHKAN LOGO DAN INFO PERUSAHAAN DI SINI =====
        # 1. Add logo (jika ada)
        logo_path = os.path.join(
            settings.BASE_DIR,
            "static",
            "images",
            "logo_3d.png"
        )
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=150, height=100)
            elements.append(logo)
            elements.append(Spacer(1, 10))
        
        # 2. Add company info
        company_info = Paragraph(
            "<b>Clean Up</b><br/>"
            "Berdiri sejak: 15 Februari 2025<br/>"
            "Alamat: Fatululi<br/>"
            "Telp: 081626261761 | Email: cleanup@yahoo.com",
            styles['Normal']
        )
        elements.append(company_info)
        elements.append(Spacer(1, 20))

        
        # Title
        report_titles = {
            'keuangan': 'LAPORAN KEUANGAN',
            'anggota': 'LAPORAN ANGGOTA', 
            'laporan-sampah': 'LAPORAN SAMPAH',
            'jadwal': 'LAPORAN JADWAL',
            'user-stats': 'STATISTIK PENGGUNA',
            'monthly': 'LAPORAN BULANAN',
            'dampak-lingkungan': 'LAPORAN DAMPAK LINGKUNGAN'
        }
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=12,
            alignment=1
        )
        
        elements.append(Paragraph(report_titles.get(report_type, 'LAPORAN'), title_style))
        
        # Add date and filters
        elements.append(Paragraph(f"Dibuat: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", 
                                 styles['Normal']))
        
        if filters:
            filter_text = " | ".join([f"{k}: {v}" for k, v in filters.items() if v])
            elements.append(Paragraph(f"Filter: {filter_text}", styles['Normal']))
        
        elements.append(Spacer(1, 20))
        
        # Period information from data
        if 'info' in data and 'period' in data['info']:
            elements.append(Paragraph(f"Periode: {data['info']['period']}", styles['Normal']))
            elements.append(Spacer(1, 10))
        
        # Add data based on report type
        if report_type == 'keuangan':
            elements.extend(self._create_keuangan_pdf(data))
        elif report_type == 'anggota':
            elements.extend(self._create_anggota_pdf(data))
        elif report_type == 'laporan-sampah':
            elements.extend(self._create_laporan_sampah_pdf(data))
        elif report_type == 'jadwal':
            elements.extend(self._create_jadwal_pdf(data))
        elif report_type == 'user-stats':
            elements.extend(self._create_user_stats_pdf(data))
        elif report_type == 'monthly':
            elements.extend(self._create_monthly_pdf(data))
        elif report_type == 'dampak-lingkungan':
            elements.extend(self._create_dampak_lingkungan_pdf(data))
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF value
        pdf = buffer.getvalue()
        buffer.close()
        
        # Create response
        response = HttpResponse(content_type='application/pdf')
        filename = f"laporan_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(pdf)
        
        return response
    
    def _format_number(self, value):
        """Safely format number, handle None and non-numeric values"""
        if value is None:
            return 0
        try:
            # Convert to float first, then to int if no decimal
            num = float(value)
            if num.is_integer():
                return int(num)
            return num
        except (ValueError, TypeError):
            return 0

    def format_date_for_report(self, date_value):
        """
        Format date for report output.
        Accepts string, datetime, date, or None.
        Returns formatted string or empty string.
        """
        if not date_value:
            return ''
        
        try:
            if isinstance(date_value, str):
                # Try to parse ISO format
                try:
                    # Handle ISO format with timezone
                    date_str = date_value.replace('Z', '+00:00')
                    dt = datetime.fromisoformat(date_str)
                except:
                    # Try other common formats
                    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S']:
                        try:
                            dt = datetime.strptime(date_value[:19], fmt)
                            break
                        except:
                            continue
                    else:
                        # Return first 10 chars if parsing fails
                        return str(date_value)[:10]
            elif hasattr(date_value, 'strftime'):
                dt = date_value
            else:
                return str(date_value)[:10]
            
            return dt.strftime('%d-%m-%Y')
        except Exception:
            return str(date_value)[:10]

    
    
    def generate_excel_response(self, data, report_type, filters):
        """Generate Excel response"""
        wb = Workbook()
        ws = wb.active
        ws.title = report_type.capitalize()[:31]  # Excel sheet name limit

        # Styling definitions
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        center_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        current_row = 1  # Start from row 1

        # ===== TAMBAHKAN LOGO DAN INFO PERUSAHAAN =====
        logo_path = os.path.join(
            settings.BASE_DIR,
            "static",
            "images",
            "logo_3d.png"
        )

        if os.path.exists(logo_path):
            # Excel tidak bisa langsung Image dari PIL/OpenPyXL dengan sizing otomatis,
            # tapi openpyxl punya Image class
            from openpyxl.drawing.image import Image as XLImage
            logo = XLImage(logo_path)
            logo.width = 200  # sesuaikan
            logo.height = 100
            ws.add_image(logo, f"A{current_row}")
            current_row += 6  # kasih spasi setelah logo

        # Company info
        company_lines = [
            "Clean Up",
            "Berdiri sejak: 15 Februari 2025",
            "Alamat: Fatululi",
            "Telp: 081626261761 | Email: cleanup@yahoo.com"
        ]
        for line in company_lines:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            cell = ws.cell(row=current_row, column=1)
            cell.value = line
            cell.font = Font(bold=True if current_row == 1 else False, color="2C3E50")
            cell.alignment = center_alignment
            current_row += 1

        current_row += 1  # spasi sebelum title laporan

        # Add title
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.value = f"LAPORAN {report_type.upper()}"
        title_cell.font = Font(bold=True, size=14, color="2C3E50")
        title_cell.alignment = center_alignment
        current_row += 1

        # Add timestamp
        ws['A' + str(current_row)] = f"Dibuat: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        current_row += 1

        # Add filters if any
        if filters:
            filter_text = " | ".join([f"{k}: {v}" for k, v in filters.items() if v])
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            cell = ws.cell(row=current_row, column=1)
            cell.value = f"Filter: {filter_text}"
            cell.alignment = center_alignment
            current_row += 2  # spasi sebelum data

        # Add data based on report type
        if report_type == 'keuangan':
            current_row = self._create_keuangan_excel(ws, data, current_row)
        elif report_type == 'anggota':
            current_row = self._create_anggota_excel(ws, data, current_row)
        elif report_type == 'laporan-sampah':
            current_row = self._create_laporan_sampah_excel(ws, data, current_row)
        elif report_type == 'jadwal':
            current_row = self._create_jadwal_excel(ws, data, current_row)
        elif report_type == 'user-stats':
            current_row = self._create_user_stats_excel(ws, data, current_row)
        elif report_type == 'monthly':
            current_row = self._create_monthly_excel(ws, data, current_row)
        elif report_type == 'dampak-lingkungan':
            current_row = self._create_dampak_lingkungan_excel(ws, data, current_row)

        # Auto adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Create response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"laporan_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def _create_keuangan_pdf(self, data):
        elements = []
        styles = getSampleStyleSheet()

        info = data.get("info", {})
        transaksi_list = data.get("table", [])  # ✅ LIST TRANSAKSI

        # =====================
        # TITLE
        # =====================
        elements.append(Paragraph("SUMMARY KEUANGAN", styles['Heading2']))
        elements.append(Spacer(1, 12))

        # =====================
        # SUMMARY
        # =====================
        summary_data = [
            ['Total Pendapatan', f"Rp {self._format_number(info.get('total_pendapatan', 0)):,.0f}"],
            ['Transaksi Pending', str(self._format_number(info.get('total_pending', 0)))],
            ['Transaksi Lunas', str(self._format_number(info.get('total_lunas', 0)))],
            ['Transaksi Gagal', str(self._format_number(info.get('total_gagal', 0)))]
        ]

        summary_table = Table(summary_data, colWidths=[200, 200])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))

        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # =====================
        # METODE BAYAR STATS
        # =====================
        metode_stats = info.get("metode_bayar_stats", {})
        if metode_stats:
            elements.append(Paragraph("Statistik Metode Pembayaran", styles["Heading3"]))
            elements.append(Spacer(1, 10))

            stats_rows = [['Metode', 'Jumlah Transaksi', 'Total (Rp)']]
            for metode, stats in metode_stats.items():
                count = stats.get("count", 0) if isinstance(stats, dict) else 0
                total = stats.get("total", 0) if isinstance(stats, dict) else 0

                stats_rows.append([
                    str(metode),
                    str(self._format_number(count)),
                    f"Rp {self._format_number(total):,.0f}"
                ])

            stats_table = Table(stats_rows, colWidths=[150, 120, 150])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ECC71')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
            ]))

            elements.append(stats_table)
            elements.append(Spacer(1, 20))

        # =====================
        # DETAIL TRANSAKSI
        # =====================
        if transaksi_list:
            elements.append(Paragraph("Detail Transaksi", styles["Heading3"]))
            elements.append(Spacer(1, 10))

            detail_rows = [
                ['No', 'Tanggal', 'Nama Anggota', 'Metode', 'Status', 'Jumlah (Rp)']
            ]

            for i, transaksi in enumerate(transaksi_list, start=1):
                detail_rows.append([
                    i,
                    self.format_date_for_report(transaksi.get('tanggal_bayar')),
                    transaksi.get('nama_anggota', '')[:20],
                    transaksi.get('metode_bayar', ''),
                    transaksi.get('status_bayar', ''),
                    f"Rp {self._format_number(transaksi.get('jumlah_bayar', 0)):,.0f}"
                ])

            detail_table = Table(detail_rows, colWidths=[30, 65, 90, 65, 55, 90])
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1),
                [colors.white, colors.HexColor('#F8F9FA')])
            ]))

            elements.append(detail_table)
            # ===== TANDA TANGAN =====
            elements.append(Spacer(1, 40))
            
            # Buat style khusus untuk nama yang bold
            bold_style = ParagraphStyle(
                'BoldStyle',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=12,
                spaceAfter=6,
                alignment=1  # Center alignment
            )
            
            normal_style = ParagraphStyle(
                'NormalStyle',
                parent=styles['Normal'],
                fontSize=12,
                spaceAfter=6,
                alignment=1  # Center alignment
            )
            
            # Tanda tangan dengan Table
            tanda_tangan_data = [
                ['', Paragraph('Hormat kami,', normal_style)],
                ['', ''],  # Baris kosong
                ['', ''],  # Baris kosong
                ['', ''],  # Baris kosong
                ['', Paragraph('<b>Julius Yohanes Belo</b>', bold_style)],  # Nama dengan bold
                ['', Paragraph('Direktur Utama', normal_style)]
            ]
            
            tanda_tangan = Table(tanda_tangan_data, colWidths=[300, 200])
            
            tanda_tangan.setStyle(TableStyle([
                ('ALIGN', (1, 4), (1, 7), 'CENTER'),  # Kolom 2, baris 4-7 center
                ('VALIGN', (1, 4), (1, 7), 'MIDDLE'),
                # Add border untuk garis tanda tangan
                ('LINEABOVE', (1, 1), (1, 3), 1, colors.black),
            ]))
            
            elements.append(tanda_tangan)

        return elements


    def _create_keuangan_excel(self, ws, data, start_row):
        info = data.get("info", {})
        table = data.get("table", [])

        # Title
        ws.merge_cells(f'A{start_row}:F{start_row}')
        title_cell = ws[f'A{start_row}']
        title_cell.value = "SUMMARY KEUANGAN"
        title_cell.font = Font(bold=True, size=12, color="2C3E50")
        title_cell.alignment = center_alignment
        
        current_row = start_row + 2
        
        # Summary
        summary_rows = [
            ("Total Pendapatan", info.get("total_pendapatan", 0)),
            ("Transaksi Pending", info.get("total_pending", 0)),
            ("Transaksi Lunas", info.get("total_lunas", 0)),
            ("Transaksi Gagal", info.get("total_gagal", 0)),
        ]

        for i, (label, val) in enumerate(summary_rows, start=current_row):
            ws[f"A{i}"] = label
            ws[f"A{i}"].font = Font(bold=True)
            ws[f"B{i}"] = self._format_number(val)
            if label == "Total Pendapatan":
                ws[f"B{i}"].number_format = '"Rp"#,##0'
        
        current_row += len(summary_rows) + 2

        # Metode Bayar Stats
        metode_stats = info.get("metode_bayar_stats", {})
        if metode_stats:
            ws[f"A{current_row}"] = "STATISTIK METODE PEMBAYARAN"
            ws[f"A{current_row}"].font = Font(bold=True, size=12, color="2C3E50")
            current_row += 2
            
            # Header
            headers = ["Metode", "Jumlah Transaksi", "Total (Rp)"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            current_row += 1
            
            # Data
            for metode, stats in metode_stats.items():
                if isinstance(stats, dict):
                    count = stats.get("count", 0)
                    total = stats.get("total", 0)
                else:
                    count = 0
                    total = 0
                
                ws.cell(row=current_row, column=1).value = str(metode)
                ws.cell(row=current_row, column=2).value = self._format_number(count)
                ws.cell(row=current_row, column=3).value = self._format_number(total)
                ws.cell(row=current_row, column=3).number_format = '"Rp"#,##0'
                current_row += 1
            
            current_row += 2

        # Detail Transaksi
        if table:
            ws[f"A{current_row}"] = "DETAIL TRANSAKSI"
            ws[f"A{current_row}"].font = Font(bold=True, size=12, color="2C3E50")
            current_row += 2
            
            # Header
            headers = ["No", "Tanggal", "Nama Anggota", "Metode", "Status", "Jumlah (Rp)"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = PatternFill(start_color="2C3E50", fill_type="solid")
                cell.alignment = center_alignment
            
            current_row += 1
            
            # Data
            for i, transaksi in enumerate(table, start=1):
                ws.cell(row=current_row, column=1).value = i
                ws.cell(row=current_row, column=2).value = self.format_date_for_report(transaksi.get('tanggal_bayar'))
                ws.cell(row=current_row, column=3).value = transaksi.get('nama_anggota', '')
                ws.cell(row=current_row, column=4).value = transaksi.get('metode_bayar', '')
                ws.cell(row=current_row, column=5).value = transaksi.get('status_bayar', '')
                
                # Format jumlah dengan Rupiah
                jumlah = transaksi.get('jumlah_bayar', 0)
                ws.cell(row=current_row, column=6).value = self._format_number(jumlah)
                ws.cell(row=current_row, column=6).number_format = '"Rp"#,##0'
                
                current_row += 1
        
        return current_row
    
    def _create_anggota_pdf(self, data):
        elements = []
        styles = getSampleStyleSheet()

        info = data.get('info', {})
        anggota_list = data.get('table', [])  # ✅ LIST DATA

        # =====================
        # TITLE
        # =====================
        elements.append(Paragraph("SUMMARY ANGGOTA", styles['Heading2']))
        elements.append(Spacer(1, 10))

        # =====================
        # SUMMARY
        # =====================
        summary_data = [
            ['Total Anggota', self._format_number(info.get('total_anggota', 0))],
            ['Aktif', self._format_number(info.get('aktif', 0))],
            ['Non Aktif', self._format_number(info.get('non_aktif', 0))],
            ['Akan Expired', self._format_number(info.get('akan_expired', 0))],
            ['Baru Bulan Ini', self._format_number(info.get('baru_bulan_ini', 0))],
        ]

        summary_table = Table(summary_data, colWidths=[200, 200])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9B59B6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))

        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # =====================
        # JENIS SAMPAH STATS
        # =====================
        jenis_stats = info.get("jenis_sampah_stats", {})
        total_anggota = info.get('total_anggota', 1) or 1

        if jenis_stats:
            elements.append(Paragraph("Distribusi Jenis Sampah", styles["Heading3"]))
            elements.append(Spacer(1, 10))

            rows = [['Jenis Sampah', 'Jumlah', 'Persentase']]
            for jenis, count in jenis_stats.items():
                count_int = int(count or 0)
                percentage = (count_int / total_anggota * 100)

                rows.append([
                    str(jenis)[:20],
                    count_int,
                    f"{percentage:.1f}%"
                ])

            stats_table = Table(rows, colWidths=[150, 100, 100])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E67E22')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
            ]))

            elements.append(stats_table)
            elements.append(Spacer(1, 20))

        # =====================
        # DETAIL ANGGOTA
        # =====================
        if anggota_list:
            elements.append(Paragraph("Detail Anggota", styles["Heading3"]))
            elements.append(Spacer(1, 10))

            detail_rows = [['No', 'Nama', 'Status', 'Jenis Sampah', 'Mulai', 'Berakhir']]
            for i, anggota in enumerate(anggota_list, start=1):
                detail_rows.append([
                    i,
                    anggota.get('nama_anggota', '')[:15],
                    anggota.get('status', ''),
                    anggota.get('jenis_sampah', '')[:10],
                    self.format_date_for_report(anggota.get('tanggal_start')),
                    self.format_date_for_report(anggota.get('tanggal_end'))
                ])

            detail_table = Table(detail_rows, colWidths=[30, 80, 50, 60, 60, 60])
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1),
                [colors.white, colors.HexColor('#F8F9FA')])
            ]))

            elements.append(detail_table)
            # ===== TANDA TANGAN =====
            elements.append(Spacer(1, 40))
            
            # Buat style khusus untuk nama yang bold
            bold_style = ParagraphStyle(
                'BoldStyle',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=12,
                spaceAfter=6,
                alignment=1  # Center alignment
            )
            
            normal_style = ParagraphStyle(
                'NormalStyle',
                parent=styles['Normal'],
                fontSize=12,
                spaceAfter=6,
                alignment=1  # Center alignment
            )
            
            # Tanda tangan dengan Table
            tanda_tangan_data = [
                ['', Paragraph('Hormat kami,', normal_style)],
                ['', ''],  # Baris kosong
                ['', ''],  # Baris kosong
                ['', ''],  # Baris kosong
                ['', Paragraph('<b>Julius Yohanes Belo</b>', bold_style)],  # Nama dengan bold
                ['', Paragraph('Direktur Utama', normal_style)]
            ]
            
            tanda_tangan = Table(tanda_tangan_data, colWidths=[300, 200])
            
            tanda_tangan.setStyle(TableStyle([
                ('ALIGN', (1, 4), (1, 7), 'CENTER'),  # Kolom 2, baris 4-7 center
                ('VALIGN', (1, 4), (1, 7), 'MIDDLE'),
                # Add border untuk garis tanda tangan
                ('LINEABOVE', (1, 1), (1, 3), 1, colors.black),
            ]))
            
            elements.append(tanda_tangan)
            
        return elements

    def _create_anggota_excel(self, ws, data, start_row):
        info = data.get('info', {})
        table = data.get('table', [])

        # Title
        ws.merge_cells(f'A{start_row}:F{start_row}')
        title_cell = ws[f'A{start_row}']
        title_cell.value = "SUMMARY ANGGOTA"
        title_cell.font = Font(bold=True, size=12, color="2C3E50")
        title_cell.alignment = center_alignment
        
        current_row = start_row + 2
        
        # Summary
        summary = [
            ('Total Anggota', info.get('total_anggota', 0)),
            ('Aktif', info.get('aktif', 0)),
            ('Non Aktif', info.get('non_aktif', 0)),
            ('Akan Expired', info.get('akan_expired', 0)),
            ('Baru Bulan Ini', info.get('baru_bulan_ini', 0)),
        ]

        for i, (label, val) in enumerate(summary, start=current_row):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = self._format_number(val)
            ws[f'A{i}'].font = Font(bold=True)
        
        current_row += len(summary) + 2

        # Jenis Sampah Stats
        jenis_stats = info.get("jenis_sampah_stats", {})
        if jenis_stats:
            ws[f'A{current_row}'] = "DISTRIBUSI JENIS SAMPAH"
            ws[f'A{current_row}'].font = Font(bold=True, size=12, color="2C3E50")
            current_row += 2
            
            # Header
            headers = ["Jenis Sampah", "Jumlah", "Persentase"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = PatternFill(start_color="27AE60", fill_type="solid")
                cell.alignment = center_alignment
            
            current_row += 1
            
            # Data
            total_anggota = info.get('total_anggota', 1)
            for jenis, count in jenis_stats.items():
                count_num = self._format_number(count)
                percentage = (count_num / total_anggota * 100) if total_anggota else 0
                
                ws.cell(row=current_row, column=1).value = str(jenis)
                ws.cell(row=current_row, column=2).value = count_num
                ws.cell(row=current_row, column=3).value = f"{percentage:.1f}%"
                current_row += 1
            
            current_row += 2

        # Detail Anggota
        if table:
            ws[f'A{current_row}'] = "DETAIL ANGGOTA"
            ws[f'A{current_row}'].font = Font(bold=True, size=12, color="2C3E50")
            current_row += 2
            
            # Header
            headers = ['No', 'Nama Anggota', 'Status', 'Jenis Sampah', 'Tanggal Mulai', 'Tanggal Berakhir']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = PatternFill(start_color="3498DB", fill_type="solid")
                cell.alignment = center_alignment
            
            current_row += 1
            
            # Data
            for i, anggota in enumerate(table, start=1):
                ws.cell(row=current_row, column=1).value = i
                ws.cell(row=current_row, column=2).value = anggota.get('nama_anggota', '')
                ws.cell(row=current_row, column=3).value = anggota.get('status', '')
                ws.cell(row=current_row, column=4).value = anggota.get('jenis_sampah', '')
                ws.cell(row=current_row, column=5).value = self.format_date_for_report(anggota.get('tanggal_start'))
                ws.cell(row=current_row, column=6).value = self.format_date_for_report(anggota.get('tanggal_end'))
                current_row += 1
        
        return current_row
    
    
    def _create_laporan_sampah_pdf(self, data):
        elements = []
        styles = getSampleStyleSheet()

        info = data.get("info", {})
        table = data.get("table", [])

        # Title
        elements.append(Paragraph("SUMMARY LAPORAN SAMPAH", styles['Heading2']))
        elements.append(Spacer(1, 10))

        # Summary data
        total = self._format_number(info.get('total_laporan', 0))
        pending = self._format_number(info.get('pending', 0))
        proses = self._format_number(info.get('proses', 0))
        selesai = self._format_number(info.get('selesai', 0))

        summary_data = [
            ['Total Laporan', total],
            ['Pending', pending],
            ['Proses', proses],
            ['Selesai', selesai],
        ]

        # Add average response time if available
        avg_response_time = info.get("avg_response_time")
        if avg_response_time is not None:
            try:
                avg_text = f"{float(avg_response_time):.1f} hari"
                summary_data.append(['Rata-rata Waktu Respons', avg_text])
            except:
                summary_data.append(['Rata-rata Waktu Respons', str(avg_response_time)])

        summary_table = Table(summary_data, colWidths=[200, 200])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E74C3C')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))

        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Detail laporan
        if table:
            elements.append(Paragraph("Detail Laporan Sampah", styles["Heading3"]))
            elements.append(Spacer(1, 10))

            table_data = [['No', 'Tanggal', 'Nama Pelapor', 'Alamat', 'Status']]
            for i, laporan in enumerate(table, start=1):
                table_data.append([
                    str(i),
                    self.format_date_for_report(laporan.get('tanggal_lapor')),
                    laporan.get('nama_pelapor', '')[:15],  # Limit nama
                    laporan.get('alamat', '')[:20],  # Limit alamat
                    laporan.get('status', '')
                ])

            detail_table = Table(table_data, colWidths=[30, 60, 70, 100, 40])
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
            ]))

            elements.append(detail_table)
            # ===== TANDA TANGAN =====
            elements.append(Spacer(1, 40))
            
            # Buat style khusus untuk nama yang bold
            bold_style = ParagraphStyle(
                'BoldStyle',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=12,
                spaceAfter=6,
                alignment=1  # Center alignment
            )
            
            normal_style = ParagraphStyle(
                'NormalStyle',
                parent=styles['Normal'],
                fontSize=12,
                spaceAfter=6,
                alignment=1  # Center alignment
            )
            
            # Tanda tangan dengan Table
            tanda_tangan_data = [
                ['', Paragraph('Hormat kami,', normal_style)],
                ['', ''],  # Baris kosong
                ['', ''],  # Baris kosong
                ['', ''],  # Baris kosong
                ['', Paragraph('<b>Julius Yohanes Belo</b>', bold_style)],  # Nama dengan bold
                ['', Paragraph('Direktur Utama', normal_style)]
            ]
            
            tanda_tangan = Table(tanda_tangan_data, colWidths=[300, 200])
            
            tanda_tangan.setStyle(TableStyle([
                ('ALIGN', (1, 4), (1, 7), 'CENTER'),  # Kolom 2, baris 4-7 center
                ('VALIGN', (1, 4), (1, 7), 'MIDDLE'),
                # Add border untuk garis tanda tangan
                ('LINEABOVE', (1, 1), (1, 3), 1, colors.black),
            ]))
            
            elements.append(tanda_tangan)

        return elements

    def _create_laporan_sampah_excel(self, ws, data, start_row):
        info = data.get("info", {})
        table = data.get("table", [])

        # Title
        ws.merge_cells(f'A{start_row}:F{start_row}')
        title_cell = ws[f'A{start_row}']
        title_cell.value = "SUMMARY LAPORAN SAMPAH"
        title_cell.font = Font(bold=True, size=12, color="2C3E50")
        title_cell.alignment = center_alignment
        
        current_row = start_row + 2
        
        # Summary
        rows = [
            ("Total Laporan", info.get("total_laporan", 0)),
            ("Pending", info.get("pending", 0)),
            ("Proses", info.get("proses", 0)),
            ("Selesai", info.get("selesai", 0)),
        ]

        for i, (label, val) in enumerate(rows, start=current_row):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = self._format_number(val)
            ws[f"A{i}"].font = Font(bold=True)
        
        current_row += len(rows) + 2

        # Add average response time if available
        avg_response_time = info.get("avg_response_time")
        if avg_response_time is not None:
            ws[f"A{current_row}"] = "Rata-rata Waktu Respons"
            ws[f"B{current_row}"] = str(avg_response_time)
            ws[f"A{current_row}"].font = Font(bold=True)
            current_row += 2

        # Detail Laporan
        if table:
            ws[f"A{current_row}"] = "DETAIL LAPORAN SAMPAH"
            ws[f"A{current_row}"].font = Font(bold=True, size=12, color="2C3E50")
            current_row += 2
            
            # Header
            headers = ["No", "Tanggal", "Nama Pelapor", "Alamat", "Status"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = PatternFill(start_color="E74C3C", fill_type="solid")
                cell.alignment = center_alignment
            
            current_row += 1
            
            # Data
            for i, laporan in enumerate(table, start=1):
                ws.cell(row=current_row, column=1).value = i
                ws.cell(row=current_row, column=2).value = self.format_date_for_report(laporan.get('tanggal_lapor'))
                ws.cell(row=current_row, column=3).value = laporan.get('nama', '')
                ws.cell(row=current_row, column=4).value = laporan.get('alamat', '')
                ws.cell(row=current_row, column=5).value = laporan.get('status', '')
                current_row += 1
        
        return current_row
    
    
    def _create_jadwal_pdf(self, data):
        elements = []
        styles = getSampleStyleSheet()

        info = data.get("info", {})
        tim_list = data.get("table", [])  # ✅ LIST DATA PER TIM

        # =====================
        # TITLE
        # =====================
        elements.append(Paragraph("SUMMARY JADWAL PENGANGKUTAN", styles['Heading2']))
        elements.append(Spacer(1, 10))

        # =====================
        # SUMMARY
        # =====================
        summary_data = [
            ['Total Jadwal', self._format_number(info.get('total_jadwal', 0))],
            ['Tim Pengangkut', self._format_number(info.get('total_tim', 0))],
            ['Anggota Terjadwal', self._format_number(info.get('total_anggota_terjadwal', 0))]
        ]

        summary_table = Table(summary_data, colWidths=[200, 200])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1ABC9C')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))

        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # =====================
        # STATUS STATS
        # =====================
        status_stats = info.get("status_stats", {})
        total_anggota = int(info.get("total_anggota_terjadwal", 1) or 1)

        if status_stats:
            elements.append(Paragraph("Status Pengangkutan", styles["Heading3"]))
            elements.append(Spacer(1, 10))

            rows = [['Status', 'Jumlah', 'Persentase']]
            for status, count in status_stats.items():
                count_int = int(count or 0)
                percent = (count_int / total_anggota) * 100

                rows.append([
                    str(status),
                    count_int,
                    f"{percent:.1f}%"
                ])

            status_table = Table(rows, colWidths=[150, 100, 100])
            status_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
            ]))

            elements.append(status_table)
            elements.append(Spacer(1, 20))

        # =====================
        # DETAIL PER TIM
        # =====================
        if tim_list:
            elements.append(Paragraph("Detail per Tim", styles["Heading3"]))
            elements.append(Spacer(1, 10))

            for tim_index, tim_data in enumerate(tim_list, start=1):
                tim_name = tim_data.get('nama_tim', f'Tim {tim_index}')
                total_jadwal = self._format_number(tim_data.get('total_jadwal', 0))
                total_anggota_tim = self._format_number(tim_data.get('total_anggota', 0))

                elements.append(Paragraph(
                    f"{tim_index}. {tim_name} (Jadwal: {total_jadwal}, Anggota: {total_anggota_tim})",
                    styles["Heading4"]
                ))

                # Status per tim
                tim_status_stats = tim_data.get('status_stats', {})
                if tim_status_stats:
                    status_rows = [['Status', 'Jumlah']]
                    for status, count in tim_status_stats.items():
                        status_rows.append([status, int(count or 0)])

                    tim_status_table = Table(status_rows, colWidths=[100, 80])
                    tim_status_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ]))
                    elements.append(tim_status_table)

                # Detail anggota
                detail_list = tim_data.get('detail', [])
                if detail_list:
                    detail_rows = [['No', 'Tanggal', 'Nama Anggota', 'Status']]
                    for i, detail in enumerate(detail_list[:5], start=1):
                        detail_rows.append([
                            i,
                            self.format_date_for_report(detail.get('tanggal_jadwal')),
                            detail.get('nama_anggota', '')[:15],
                            detail.get('status_pengangkutan', '')
                        ])

                    detail_table = Table(detail_rows, colWidths=[25, 60, 80, 60])
                    detail_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ]))
                    elements.append(detail_table)

                if len(detail_list) > 5:
                    elements.append(
                        Paragraph(f"... dan {len(detail_list) - 5} anggota lainnya", styles['Italic'])
                    )

                elements.append(Spacer(1, 12))
                # ===== TANDA TANGAN =====
                elements.append(Spacer(1, 40))
                
                # Buat style khusus untuk nama yang bold
                bold_style = ParagraphStyle(
                    'BoldStyle',
                    parent=styles['Normal'],
                    fontName='Helvetica-Bold',
                    fontSize=12,
                    spaceAfter=6,
                    alignment=1  # Center alignment
                )
                
                normal_style = ParagraphStyle(
                    'NormalStyle',
                    parent=styles['Normal'],
                    fontSize=12,
                    spaceAfter=6,
                    alignment=1  # Center alignment
                )
                
                # Tanda tangan dengan Table
                tanda_tangan_data = [
                    ['', Paragraph('Hormat kami,', normal_style)],
                    ['', ''],  # Baris kosong
                    ['', ''],  # Baris kosong
                    ['', ''],  # Baris kosong
                    ['', Paragraph('<b>Julius Yohanes Belo</b>', bold_style)],  # Nama dengan bold
                    ['', Paragraph('Direktur Utama', normal_style)]
                ]
                
                tanda_tangan = Table(tanda_tangan_data, colWidths=[300, 200])
                
                tanda_tangan.setStyle(TableStyle([
                    ('ALIGN', (1, 4), (1, 7), 'CENTER'),  # Kolom 2, baris 4-7 center
                    ('VALIGN', (1, 4), (1, 7), 'MIDDLE'),
                    # Add border untuk garis tanda tangan
                    ('LINEABOVE', (1, 1), (1, 3), 1, colors.black),
                ]))
                
                elements.append(tanda_tangan)
        return elements


    def _create_jadwal_excel(self, ws, data, start_row):
        info = data.get("info", {})
        table = data.get("table", [])

        # Title
        ws.merge_cells(f'A{start_row}:F{start_row}')
        title_cell = ws[f'A{start_row}']
        title_cell.value = "SUMMARY JADWAL PENGANGKUTAN"
        title_cell.font = Font(bold=True, size=12, color="2C3E50")
        title_cell.alignment = center_alignment
        
        current_row = start_row + 2
        
        # Summary
        rows = [
            ("Total Jadwal", info.get("total_jadwal", 0)),
            ("Tim Pengangkut", info.get("total_tim", 0)),
            ("Anggota Terjadwal", info.get("total_anggota_terjadwal", 0)),
        ]

        for i, (label, val) in enumerate(rows, start=current_row):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = self._format_number(val)
            ws[f"A{i}"].font = Font(bold=True)
        
        current_row += len(rows) + 2

        # Status Stats
        status_stats = info.get("status_stats", {})
        if status_stats:
            ws[f"A{current_row}"] = "STATUS PENGANGKUTAN"
            ws[f"A{current_row}"].font = Font(bold=True, size=12, color="2C3E50")
            current_row += 2
            
            # Header
            headers = ["Status", "Jumlah", "Persentase"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = PatternFill(start_color="9B59B6", fill_type="solid")
                cell.alignment = center_alignment
            
            current_row += 1
            
            # Data
            total = info.get("total_anggota_terjadwal", 1)
            for status, count in status_stats.items():
                count_num = self._format_number(count)
                percentage = (count_num / total * 100) if total else 0
                
                ws.cell(row=current_row, column=1).value = str(status)
                ws.cell(row=current_row, column=2).value = count_num
                ws.cell(row=current_row, column=3).value = f"{percentage:.1f}%"
                current_row += 1
            
            current_row += 2

        # Detail per Tim
        if table:
            ws[f"A{current_row}"] = "DETAIL PER TIM"
            ws[f"A{current_row}"].font = Font(bold=True, size=12, color="2C3E50")
            current_row += 2
            
            for tim_index, tim_data in enumerate(table, start=1):
                # Header per tim
                tim_name = tim_data.get('nama_tim', f'Tim {tim_index}')
                total_jadwal = self._format_number(tim_data.get('total_jadwal', 0))
                total_anggota = self._format_number(tim_data.get('total_anggota', 0))
                
                ws[f"A{current_row}"] = f"{tim_index}. {tim_name} (Jadwal: {total_jadwal}, Anggota: {total_anggota})"
                ws[f"A{current_row}"].font = Font(bold=True, color="2C3E50")
                current_row += 1
                
                # Status stats per tim
                tim_status_stats = tim_data.get('status_stats', {})
                if tim_status_stats:
                    # Sub-header
                    ws[f"B{current_row}"] = "Status"
                    ws[f"C{current_row}"] = "Jumlah"
                    ws[f"B{current_row}"].font = Font(bold=True)
                    ws[f"C{current_row}"].font = Font(bold=True)
                    current_row += 1
                    
                    for status, count in tim_status_stats.items():
                        ws[f"B{current_row}"] = str(status)
                        ws[f"C{current_row}"] = self._format_number(count)
                        current_row += 1
                
                # Detail anggota
                detail_list = tim_data.get('detail', [])
                if detail_list:
                    current_row += 1
                    ws[f"A{current_row}"] = "Detail Anggota:"
                    ws[f"A{current_row}"].font = Font(bold=True)
                    current_row += 1
                    
                    # Sub-table header
                    sub_headers = ["No", "Tanggal", "Nama Anggota", "Status"]
                    for col, header in enumerate(sub_headers, start=1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.value = header
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="34495E", fill_type="solid")
                        cell.alignment = center_alignment
                    
                    current_row += 1
                    
                    for i, detail in enumerate(detail_list, start=1):
                        ws.cell(row=current_row, column=1).value = i
                        ws.cell(row=current_row, column=2).value = self.format_date_for_report(detail.get('tanggal_jadwal'))
                        ws.cell(row=current_row, column=3).value = detail.get('nama_anggota', '')
                        ws.cell(row=current_row, column=4).value = detail.get('status_pengangkutan', '')
                        current_row += 1
                
                current_row += 2  # Spacing antar tim
        
        return current_row
       
    def _create_user_stats_pdf(self, data):
        elements = []
        styles = getSampleStyleSheet()

        info = data.get("info", {})
        table = data.get("table", [])

        # Title
        elements.append(Paragraph("SUMMARY STATISTIK PENGGUNA", styles['Heading2']))
        elements.append(Spacer(1, 10))

        # Summary data
        total_users = self._format_number(info.get('total_users', 0))
        active_users = self._format_number(info.get('active_users', 0))
        admin_count = self._format_number(info.get('admin_count', 0))
        anggota_count = self._format_number(info.get('anggota_count', 0))
        tamu_count = self._format_number(info.get('tamu_count', 0))
        tim_angkut_count = self._format_number(info.get('tim_angkut_count', 0))
        new_users_month = self._format_number(info.get('new_users_month', 0))

        summary_data = [
            ['Total Pengguna', total_users],
            ['Pengguna Aktif', f"{active_users} ({active_users/total_users*100:.1f}%)" if total_users else "0"],
            ['Admin', admin_count],
            ['Anggota', anggota_count],
            ['Tamu', tamu_count],
            ['Tim Pengangkut', tim_angkut_count],
            ['Pengguna Baru Bulan Ini', new_users_month]
        ]

        summary_table = Table(summary_data, colWidths=[200, 200])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F39C12')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))

        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Detail pengguna
        if table:
            elements.append(Paragraph("Detail Pengguna", styles["Heading3"]))
            elements.append(Spacer(1, 10))

            table_data = [['No', 'Username', 'Email', 'Role', 'Aktif', 'Tanggal Daftar']]
            for i, user in enumerate(table, start=1):
                table_data.append([
                    str(i),
                    user.get('username', '')[:15],  # Limit username
                    user.get('email', '')[:20],  # Limit email
                    user.get('role', ''),
                    'Ya' if user.get('is_active') else 'Tidak',
                    self.format_date_for_report(user.get('date_joined'))
                ])

            detail_table = Table(table_data, colWidths=[25, 60, 80, 50, 40, 60])
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
            ]))

            elements.append(detail_table)
            # ===== TANDA TANGAN =====
            elements.append(Spacer(1, 40))
            
            # Buat style khusus untuk nama yang bold
            bold_style = ParagraphStyle(
                'BoldStyle',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=12,
                spaceAfter=6,
                alignment=1  # Center alignment
            )
            
            normal_style = ParagraphStyle(
                'NormalStyle',
                parent=styles['Normal'],
                fontSize=12,
                spaceAfter=6,
                alignment=1  # Center alignment
            )
            
            # Tanda tangan dengan Table
            tanda_tangan_data = [
                ['', Paragraph('Hormat kami,', normal_style)],
                ['', ''],  # Baris kosong
                ['', ''],  # Baris kosong
                ['', ''],  # Baris kosong
                ['', Paragraph('<b>Julius Yohanes Belo</b>', bold_style)],  # Nama dengan bold
                ['', Paragraph('Direktur Utama', normal_style)]
            ]
            
            tanda_tangan = Table(tanda_tangan_data, colWidths=[300, 200])
            
            tanda_tangan.setStyle(TableStyle([
                ('ALIGN', (1, 4), (1, 7), 'CENTER'),  # Kolom 2, baris 4-7 center
                ('VALIGN', (1, 4), (1, 7), 'MIDDLE'),
                # Add border untuk garis tanda tangan
                ('LINEABOVE', (1, 1), (1, 3), 1, colors.black),
            ]))
            
            elements.append(tanda_tangan)

        return elements

    def _create_user_stats_excel(self, ws, data, start_row):
        info = data.get("info", {})
        table_data = data.get("table", [])

        # Title
        ws.merge_cells(f'A{start_row}:F{start_row}')
        title_cell = ws[f'A{start_row}']
        title_cell.value = "SUMMARY STATISTIK PENGGUNA"
        title_cell.font = Font(bold=True, size=12, color="2C3E50")
        title_cell.alignment = center_alignment
        
        current_row = start_row + 2
        
        # Summary
        summary_rows = [
            ("Keterangan", "Jumlah"),
            ("Total Pengguna", info.get("total_users", 0)),
            ("Pengguna Aktif", info.get("active_users", 0)),
            ("Admin", info.get("admin_count", 0)),
            ("Anggota", info.get("anggota_count", 0)),
            ("Tamu", info.get("tamu_count", 0)),
            ("Tim Pengangkut", info.get("tim_angkut_count", 0)),
            ("Pengguna Baru Bulan Ini", info.get("new_users_month", 0)),
        ]

        for i, (label, val) in enumerate(summary_rows, start=current_row):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = self._format_number(val) if i > current_row else val
            ws[f"A{i}"].font = Font(bold=True) if i == current_row else Font()
        
        current_row += len(summary_rows) + 2

        # Detail Pengguna
        if table_data:
            ws[f"A{current_row}"] = "DETAIL PENGGUNA"
            ws[f"A{current_row}"].font = Font(bold=True, size=12, color="2C3E50")
            current_row += 2
            
            # Header
            headers = ["No", "Username", "Email", "Role", "Aktif", "Tanggal Daftar"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = PatternFill(start_color="2C3E50", fill_type="solid")
                cell.alignment = center_alignment
            
            current_row += 1
            
            # Data
            for i, row in enumerate(table_data, start=1):
                ws.cell(row=current_row, column=1).value = i
                ws.cell(row=current_row, column=2).value = row.get("username", "")
                ws.cell(row=current_row, column=3).value = row.get("email", "")
                ws.cell(row=current_row, column=4).value = row.get("role", "")
                ws.cell(row=current_row, column=5).value = "Ya" if row.get("is_active") else "Tidak"
                ws.cell(row=current_row, column=6).value = self.format_date_for_report(row.get("date_joined"))
                current_row += 1
        
        return current_row
    
    def _create_monthly_pdf(self, data):
        """Create PDF content for monthly report"""
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        elements.append(Paragraph(data.get('bulan', 'Laporan Bulanan'), styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        # Metrics
        metrics_data = [
            ['Metrik', 'Nilai'],
            ['Pendapatan', f"Rp {self._format_number(data.get('total_pendapatan', 0)):,.0f}"],
            ['Transaksi', str(self._format_number(data.get('total_transaksi', 0)))],
            ['Anggota', str(self._format_number(data.get('total_anggota', 0)))],
            ['Anggota Baru', str(self._format_number(data.get('anggota_baru', 0)))],
            ['Jadwal', str(self._format_number(data.get('total_jadwal', 0)))],
            ['Anggota Dilayani', str(self._format_number(data.get('anggota_dilayani', 0)))],
            ['Success Rate', f"{self._format_number(data.get('success_rate', 0)):.1f}%"],
            ['Laporan Sampah', str(self._format_number(data.get('total_laporan', 0)))],
            ['Resolution Rate', f"{self._format_number(data.get('resolution_rate', 0)):.1f}%"]
        ]
        
        metrics_table = Table(metrics_data, colWidths=[200, 200])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))
        
        elements.append(metrics_table)
        elements.append(Spacer(1, 20))
        
        # Summary jika ada
        summary = data.get('summary', {})
        if summary:
            elements.append(Paragraph("Analisis Performance", styles["Heading3"]))
            elements.append(Spacer(1, 10))
            
            analysis_data = [
                ['Pendapatan per Anggota', f"Rp {self._format_number(summary.get('pendapatan_per_anggota', 0)):,.0f}"],
                ['Laporan per User', f"{self._format_number(summary.get('laporan_per_user', 0)):.2f}"],
                ['Efficiency Rate', f"{self._format_number(summary.get('efficiency_rate', 0)):.1f}%"]
            ]
            
            analysis_table = Table(analysis_data, colWidths=[200, 200])
            analysis_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            elements.append(analysis_table)
            # ===== TANDA TANGAN =====
            elements.append(Spacer(1, 40))
            
            # Buat style khusus untuk nama yang bold
            bold_style = ParagraphStyle(
                'BoldStyle',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=12,
                spaceAfter=6,
                alignment=1  # Center alignment
            )
            
            normal_style = ParagraphStyle(
                'NormalStyle',
                parent=styles['Normal'],
                fontSize=12,
                spaceAfter=6,
                alignment=1  # Center alignment
            )
            
            # Tanda tangan dengan Table
            tanda_tangan_data = [
                ['', Paragraph('Hormat kami,', normal_style)],
                ['', ''],  # Baris kosong
                ['', ''],  # Baris kosong
                ['', ''],  # Baris kosong
                ['', Paragraph('<b>Julius Yohanes Belo</b>', bold_style)],  # Nama dengan bold
                ['', Paragraph('Direktur Utama', normal_style)]
            ]
            
            tanda_tangan = Table(tanda_tangan_data, colWidths=[300, 200])
            
            tanda_tangan.setStyle(TableStyle([
                ('ALIGN', (1, 4), (1, 7), 'CENTER'),  # Kolom 2, baris 4-7 center
                ('VALIGN', (1, 4), (1, 7), 'MIDDLE'),
                # Add border untuk garis tanda tangan
                ('LINEABOVE', (1, 1), (1, 3), 1, colors.black),
            ]))
            
            elements.append(tanda_tangan)
        
        return elements

    def _create_monthly_excel(self, ws, data, start_row):
        # Title
        ws.merge_cells(f'A{start_row}:F{start_row}')
        title_cell = ws[f'A{start_row}']
        title_cell.value = f"LAPORAN BULANAN - {data.get('bulan', '')}"
        title_cell.font = Font(bold=True, size=12, color="2C3E50")
        title_cell.alignment = center_alignment
        
        current_row = start_row + 2
        
        # Header
        ws[f"A{current_row}"] = "METRIK"
        ws[f"B{current_row}"] = "NILAI"
        ws[f"A{current_row}"].font = Font(bold=True, color="FFFFFF")
        ws[f"B{current_row}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{current_row}"].fill = PatternFill(start_color="3498DB", fill_type="solid")
        ws[f"B{current_row}"].fill = PatternFill(start_color="3498DB", fill_type="solid")
        ws[f"A{current_row}"].alignment = center_alignment
        ws[f"B{current_row}"].alignment = center_alignment
        
        current_row += 1
        
        # Metrics
        metrics = [
            ("Pendapatan", data.get('total_pendapatan', 0), "Rp"),
            ("Transaksi", data.get('total_transaksi', 0), ""),
            ("Anggota", data.get('total_anggota', 0), ""),
            ("Anggota Baru", data.get('anggota_baru', 0), ""),
            ("Jadwal", data.get('total_jadwal', 0), ""),
            ("Anggota Dilayani", data.get('anggota_dilayani', 0), ""),
            ("Success Rate", data.get('success_rate', 0), "%"),
            ("Laporan Sampah", data.get('total_laporan', 0), ""),
            ("Resolution Rate", data.get('resolution_rate', 0), "%"),
        ]
        
        for label, value, unit in metrics:
            ws[f"A{current_row}"] = label
            ws[f"A{current_row}"].font = Font(bold=True)
            
            if unit == "Rp":
                ws[f"B{current_row}"] = self._format_number(value)
                ws[f"B{current_row}"].number_format = '"Rp"#,##0'
            elif unit == "%":
                ws[f"B{current_row}"] = f"{self._format_number(value):.1f}{unit}"
            else:
                ws[f"B{current_row}"] = self._format_number(value)
            
            current_row += 1
        
        current_row += 2
        
        # Summary jika ada
        summary = data.get('summary', {})
        if summary:
            ws[f"A{current_row}"] = "ANALISIS PERFORMANCE"
            ws[f"A{current_row}"].font = Font(bold=True, size=12, color="2C3E50")
            current_row += 2
            
            analysis_metrics = [
                ("Pendapatan per Anggota", summary.get('pendapatan_per_anggota', 0), "Rp"),
                ("Laporan per User", summary.get('laporan_per_user', 0), ""),
                ("Efficiency Rate", summary.get('efficiency_rate', 0), "%"),
            ]
            
            for label, value, unit in analysis_metrics:
                ws[f"A{current_row}"] = label
                ws[f"A{current_row}"].font = Font(bold=True)
                
                if unit == "Rp":
                    ws[f"B{current_row}"] = self._format_number(value)
                    ws[f"B{current_row}"].number_format = '"Rp"#,##0'
                elif unit == "%":
                    ws[f"B{current_row}"] = f"{self._format_number(value):.1f}{unit}"
                else:
                    ws[f"B{current_row}"] = f"{self._format_number(value):.2f}"
                
                current_row += 1
        
        return current_row

    def _create_dampak_lingkungan_pdf(self, data):
        """Create PDF content for dampak lingkungan report"""
        elements = []
        styles = getSampleStyleSheet()
        
        # Judul Laporan
        elements.append(Paragraph('LAPORAN ANALISIS DAMPAK LINGKUNGAN', styles['Title']))
        elements.append(Spacer(1, 20))
        
        # PERBAIKAN: Summary - menggunakan data yang benar
        summary_data = [
            ['Total Laporan', str(data.get('total_laporan', 0))],
            ['Laporan Selesai', str(data.get('laporan_selesai', 0))],
            ['Tingkat Penyelesaian', f"{data.get('tingkat_penyelesaian', 0):.1f}%"]
        ]
        
        # PERBAIKAN: Akses data dampak_lingkungan yang benar
        if 'dampak_lingkungan' in data:
            dampak = data['dampak_lingkungan']
            ringkasan = dampak.get('ringkasan', {})
            
            # Tambahkan data dari ringkasan dampak lingkungan
            summary_data.append(['Jenis Berbahaya', str(ringkasan.get('total_jenis_berbahaya', 0))])
            summary_data.append(['Total Peringatan', str(ringkasan.get('total_peringatan', 0))])
            summary_data.append(['Tingkat Risiko', str(ringkasan.get('tingkat_risiko', 'aman')).capitalize()])
        
        summary_table = Table(summary_data, colWidths=[200, 150])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ECF0F1')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # PERBAIKAN: Klasifikasi sampah - format data yang benar
        if 'klasifikasi_sampah' in data and data['klasifikasi_sampah']:
            klasifikasi = data['klasifikasi_sampah']
            elements.append(Paragraph('Klasifikasi Jenis Sampah', styles['Heading2']))
            
            if 'detail_klasifikasi' in klasifikasi and klasifikasi['detail_klasifikasi']:
                ks_data = [['Jenis Sampah', 'Jumlah Laporan', 'Persentase']]
                
                for item in klasifikasi['detail_klasifikasi']:
                    if item.get('jumlah', 0) > 0:
                        ks_data.append([
                            str(item.get('jenis', 'tidak diketahui')).capitalize().replace('_', ' '),
                            str(item.get('jumlah', 0)),
                            f"{item.get('persentase', 0):.1f}%"
                        ])
                
                if len(ks_data) > 1:  # Ada data selain header
                    ks_table = Table(ks_data, colWidths=[150, 100, 100])
                    ks_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    
                    elements.append(ks_table)
                    elements.append(Spacer(1, 10))
                    
                    # Info total data
                    elements.append(Paragraph(
                        f"Total data terstruktur: {klasifikasi.get('total_data_terstruktur', 0)} "
                        f"({klasifikasi.get('persentase_data_terstruktur', 0):.1f}%)",
                        styles['Italic']
                    ))
            
            elements.append(Spacer(1, 20))
        
        # PERBAIKAN: Analisis dampak lingkungan
        if 'dampak_lingkungan' in data:
            dampak = data['dampak_lingkungan']
            elements.append(Paragraph('Analisis Dampak Lingkungan', styles['Heading2']))
            
            # Ringkasan dampak
            ringkasan = dampak.get('ringkasan', {})
            de_data = [
                ['Parameter', 'Nilai'],
                ['Total Analisis', str(dampak.get('total_analisis', 0))],
                ['Jenis Berbahaya', str(ringkasan.get('total_jenis_berbahaya', 0))],
                ['Total Peringatan', str(ringkasan.get('total_peringatan', 0))],
                ['Tingkat Risiko', str(ringkasan.get('tingkat_risiko', 'aman')).capitalize()]
            ]
            
            de_table = Table(de_data, colWidths=[200, 150])
            de_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E74C3C')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
            ]))
            
            elements.append(de_table)
            elements.append(Spacer(1, 20))
            
            # PERBAIKAN: Peringatan dampak lingkungan
            if 'peringatan' in dampak and dampak['peringatan']:
                elements.append(Paragraph('Peringatan Dampak Lingkungan', styles['Heading3']))
                
                for i, warning in enumerate(dampak['peringatan'][:5], start=1):
                    # Warna berdasarkan level peringatan
                    warna_peringatan = {
                        'sangat_tinggi': '#E74C3C',
                        'tinggi': '#E67E22',
                        'sedang': '#F1C40F',
                        'rendah': '#2ECC71',
                        'aman': '#27AE60'
                    }
                    
                    level = warning.get('level', 'sedang')
                    elements.append(Paragraph(
                        f"{i}. [{level.upper()}] {warning.get('jenis', 'tidak diketahui').capitalize()}",
                        ParagraphStyle(
                            'CustomWarning',
                            parent=styles['Normal'],
                            textColor=colors.HexColor(warna_peringatan.get(level, '#000000')),
                            fontSize=11,
                            spaceAfter=4,
                            leftIndent=10,
                            fontName='Helvetica-Bold'
                        )
                    ))
                    
                    # Dampak potensial (ambil yang pertama jika berupa list)
                    dampak_list = warning.get('dampak', [])
                    if isinstance(dampak_list, list) and dampak_list:
                        dampak_text = dampak_list[0] if dampak_list else 'Tidak tersedia'
                    else:
                        dampak_text = str(dampak_list)
                    
                    elements.append(Paragraph(
                        f"   • Jumlah: {warning.get('jumlah', 0)} ({warning.get('persentase', 0):.1f}%)",
                        styles['Italic']
                    ))
                    elements.append(Paragraph(
                        f"   • Dampak: {dampak_text[:80]}...",
                        styles['Italic']
                    ))
                    elements.append(Spacer(1, 8))
                
                elements.append(Spacer(1, 20))
        
        # PERBAIKAN: Hotspot lokasi - gunakan data yang benar
        if 'ranking_wilayah' in data and data['ranking_wilayah']:
            ranking = data['ranking_wilayah']
            elements.append(Paragraph('Wilayah dengan Tingkat Kebersihan Terendah', styles['Heading2']))
            
            if 'ranking_terkotor' in ranking and ranking['ranking_terkotor']:
                hs_data = [['No', 'Wilayah', 'Total Laporan', 'Selesai', 'Skor Kebersihan']]
                
                for i, hotspot in enumerate(ranking['ranking_terkotor'][:5], start=1):
                    hs_data.append([
                        str(i),
                        hotspot.get('wilayah', 'Tidak diketahui')[:30],
                        str(hotspot.get('total_laporan', 0)),
                        str(hotspot.get('laporan_selesai', 0)),
                        f"{hotspot.get('skor_kebersihan', 0):.1f}"
                    ])
                
                hs_table = Table(hs_data, colWidths=[40, 120, 70, 70, 80])
                hs_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9B59B6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(hs_table)
                elements.append(Spacer(1, 20))
        
        # PERBAIKAN: Rekomendasi
        if 'rekomendasi' in data and data['rekomendasi']:
            elements.append(Paragraph('Rekomendasi Tindakan', styles['Heading2']))
            
            for i, rec in enumerate(data['rekomendasi'], start=1):
                prioritas_warna = {
                    'sangat_tinggi': '#E74C3C',
                    'tinggi': '#E67E22',
                    'sedang': '#F1C40F',
                    'rendah': '#2ECC71'
                }
                
                # Style berdasarkan prioritas
                prioritas = rec.get('prioritas', 'sedang')
                text_color = colors.HexColor(prioritas_warna.get(prioritas, '#000000'))
                
                elements.append(Paragraph(
                    f"{i}. [{prioritas.upper()}] {rec['rekomendasi']}",
                    ParagraphStyle(
                        'CustomRecommendation',
                        parent=styles['Normal'],
                        textColor=text_color,
                        fontSize=11,
                        spaceAfter=4,
                        leftIndent=10,
                        fontName='Helvetica-Bold'
                    )
                ))
                
                # Informasi tambahan
                info_lines = []
                if rec.get('kategori'):
                    info_lines.append(f"Kategori: {rec.get('kategori')}")
                if rec.get('alasan'):
                    info_lines.append(f"Alasan: {rec.get('alasan')}")
                if rec.get('sumber_data'):
                    info_lines.append(f"Sumber Data: {rec.get('sumber_data')}")
                
                if info_lines:
                    elements.append(Paragraph(
                        f"   {' | '.join(info_lines)}",
                        styles['Italic']
                    ))
                
                elements.append(Spacer(1, 12))
        
        # ===== TANDA TANGAN =====
        elements.append(Spacer(1, 40))
        
        # Buat style khusus
        bold_style = ParagraphStyle(
            'BoldStyle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=12,
            spaceAfter=6,
            alignment=1  # Center alignment
        )
        
        normal_style = ParagraphStyle(
            'NormalStyle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=6,
            alignment=1
        )
        
        # Tanda tangan dengan Table
        tanda_tangan_data = [
            ['', Paragraph('Hormat kami,', normal_style)],
            ['', ''],  # Baris kosong untuk garis tanda tangan
            ['', ''],  # Baris kosong
            ['', ''],  # Baris kosong
            ['', Paragraph('<b>Julius Yohanes Belo</b>', bold_style)],
            ['', Paragraph('Direktur Utama', normal_style)]
        ]
        
        tanda_tangan = Table(tanda_tangan_data, colWidths=[300, 200])
        
        tanda_tangan.setStyle(TableStyle([
            ('ALIGN', (1, 0), (1, 5), 'CENTER'),
            ('VALIGN', (1, 0), (1, 5), 'MIDDLE'),
            ('LINEABOVE', (1, 1), (1, 1), 1, colors.black),  # Garis tanda tangan
        ]))
        
        elements.append(tanda_tangan)
        
        # Info periode dan tanggal generate
        elements.append(Spacer(1, 30))
        
        info_lines = []
        if 'period' in data and isinstance(data['period'], dict):
            period = data['period']
            if 'label' in period:
                info_lines.append(f"Periode Analisis: {period['label']}")
        
        if 'tanggal_generate' in data:
            try:
                # Format tanggal generate
                from django.utils import timezone
                if isinstance(data['tanggal_generate'], str):
                    tanggal_str = data['tanggal_generate']
                else:
                    tanggal_str = data['tanggal_generate'].strftime('%d %B %Y, %H:%M')
                info_lines.append(f"Tanggal Generate: {tanggal_str}")
            except:
                info_lines.append(f"Tanggal Generate: {str(data['tanggal_generate'])}")
        
        if info_lines:
            elements.append(Paragraph(
                " | ".join(info_lines),
                ParagraphStyle(
                    'CustomFooter',
                    parent=styles['Normal'],
                    fontSize=9,
                    textColor=colors.gray,
                    alignment=1
                )
            ))
        
        return elements

    def _create_dampak_lingkungan_excel(self, ws, data, start_row):
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Styling
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        current_row = start_row

        # ===================== JUDUL LAPORAN =====================
        ws[f'A{current_row}'] = "LAPORAN ANALISIS DAMPAK LINGKUNGAN"
        ws[f'A{current_row}'].font = Font(bold=True, size=14, color="2C3E50")
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 2

        # ===================== SUMMARY =====================
        ws[f'A{current_row}'] = "SUMMARY"
        ws[f'A{current_row}'].font = Font(bold=True, size=12, color="27AE60")
        current_row += 2

        summary_data = [
            ('Total Laporan', data.get('total_laporan', 0)),
            ('Laporan Selesai', data.get('laporan_selesai', 0)),
            ('Tingkat Penyelesaian', f"{data.get('tingkat_penyelesaian', 0):.1f}%")
        ]

        # PERBAIKAN: Tambahkan data dari dampak_lingkungan jika ada
        if 'dampak_lingkungan' in data:
            dampak = data['dampak_lingkungan']
            ringkasan = dampak.get('ringkasan', {})
            summary_data.append(('Total Analisis', dampak.get('total_analisis', 0)))
            summary_data.append(('Jenis Berbahaya', ringkasan.get('total_jenis_berbahaya', 0)))
            summary_data.append(('Total Peringatan', ringkasan.get('total_peringatan', 0)))
            summary_data.append(('Tingkat Risiko', ringkasan.get('tingkat_risiko', 'aman').capitalize()))

        for label, val in summary_data:
            ws[f'A{current_row}'] = label
            ws[f'B{current_row}'] = val
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'A{current_row}'].alignment = left_alignment
            ws[f'B{current_row}'].alignment = left_alignment
            current_row += 1

        current_row += 2

        # ===================== KLASIFIKASI SAMPAH =====================
        klasifikasi = data.get('klasifikasi_sampah', {})
        if klasifikasi and 'detail_klasifikasi' in klasifikasi and klasifikasi['detail_klasifikasi']:
            ws[f'A{current_row}'] = "KLASIFIKASI JENIS SAMPAH"
            ws[f'A{current_row}'].font = Font(bold=True, size=12, color="3498DB")
            current_row += 2

            # Header
            headers = ["Jenis Sampah", "Jumlah", "Persentase", "Status"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = PatternFill(start_color="3498DB", fill_type="solid")
                cell.alignment = center_alignment
                cell.border = border
            current_row += 1

            # Data
            for item in klasifikasi['detail_klasifikasi']:
                jumlah = item.get('jumlah', 0)
                if jumlah > 0:
                    ws.cell(row=current_row, column=1).value = str(item.get('jenis', '')).capitalize().replace('_', ' ')
                    ws.cell(row=current_row, column=2).value = jumlah
                    ws.cell(row=current_row, column=3).value = f"{item.get('persentase', 0):.1f}%"
                    ws.cell(row=current_row, column=4).value = item.get('status', 'aman')
                    
                    # Warna latar berdasarkan status
                    status_fill = {
                        'perhatian': PatternFill(start_color="FFF2CC", fill_type="solid"),
                        'bahaya': PatternFill(start_color="FFCCCC", fill_type="solid"),
                        'aman': PatternFill(start_color="D5E8D4", fill_type="solid")
                    }
                    ws.cell(row=current_row, column=4).fill = status_fill.get(
                        item.get('status', 'aman'), 
                        PatternFill(start_color="FFFFFF", fill_type="solid")
                    )
                    
                    # Border untuk semua cell
                    for col in range(1, 5):
                        ws.cell(row=current_row, column=col).border = border
                        ws.cell(row=current_row, column=col).alignment = center_alignment
                    
                    current_row += 1

            # Info total data
            current_row += 1
            ws[f'A{current_row}'] = "Informasi Data:"
            ws[f'B{current_row}'] = f"Total Data Terstruktur: {klasifikasi.get('total_data_terstruktur', 0)}"
            ws[f'C{current_row}'] = f"({klasifikasi.get('persentase_data_terstruktur', 0):.1f}%)"
            ws.merge_cells(f'A{current_row}:C{current_row}')
            current_row += 2

        # ===================== DAMPAK LINGKUNGAN =====================
        # PERBAIKAN: Gunakan data yang benar
        dampak = data.get('dampak_lingkungan')
        if dampak:
            ws[f'A{current_row}'] = "ANALISIS DAMPAK LINGKUNGAN"
            ws[f'A{current_row}'].font = Font(bold=True, size=12, color="E74C3C")
            current_row += 2

            # Ringkasan dampak
            ringkasan = dampak.get('ringkasan', {})
            dampak_items = [
                ('Total Analisis', dampak.get('total_analisis', 0), 'laporan'),
                ('Jenis Berbahaya', ringkasan.get('total_jenis_berbahaya', 0), 'jenis'),
                ('Total Peringatan', ringkasan.get('total_peringatan', 0), 'item'),
                ('Tingkat Risiko', ringkasan.get('tingkat_risiko', 'aman').capitalize(), ''),
            ]

            # Header
            headers = ['Parameter', 'Nilai', 'Satuan/Keterangan']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = PatternFill(start_color="E74C3C", fill_type="solid")
                cell.alignment = center_alignment
                cell.border = border
            current_row += 1

            # Data
            for param, nilai, satuan in dampak_items:
                ws.cell(row=current_row, column=1).value = param
                ws.cell(row=current_row, column=2).value = nilai
                ws.cell(row=current_row, column=3).value = satuan
                
                # Border untuk semua cell
                for col in range(1, 4):
                    ws.cell(row=current_row, column=col).border = border
                    ws.cell(row=current_row, column=col).alignment = left_alignment
                
                current_row += 1

            current_row += 2

            # ===================== PERINGATAN DAMPAK =====================
            peringatan = dampak.get('peringatan', [])[:10]  # Ambil 10 teratas
            if peringatan:
                ws[f'A{current_row}'] = "PERINGATAN DAMPAK LINGKUNGAN (TOP 10)"
                ws[f'A{current_row}'].font = Font(bold=True, size=12, color="FF0000")
                current_row += 2

                headers = ['No', 'Jenis Sampah', 'Level', 'Jumlah', 'Persentase', 'Dampak Utama', 'Rekomendasi']
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = PatternFill(start_color="FF0000", fill_type="solid")
                    cell.alignment = center_alignment
                    cell.border = border
                current_row += 1

                # Warna berdasarkan level
                level_colors = {
                    'sangat_tinggi': PatternFill(start_color="FF0000", fill_type="solid"),  # Merah
                    'tinggi': PatternFill(start_color="FF9900", fill_type="solid"),        # Oranye
                    'sedang': PatternFill(start_color="FFFF00", fill_type="solid"),        # Kuning
                    'rendah': PatternFill(start_color="00FF00", fill_type="solid"),        # Hijau
                    'aman': PatternFill(start_color="CCCCCC", fill_type="solid")           # Abu-abu
                }

                for i, warning in enumerate(peringatan, start=1):
                    # Data peringatan
                    jenis = warning.get('jenis', 'tidak diketahui').capitalize()
                    level = warning.get('level', 'sedang')
                    
                    # Handle dampak (bisa list atau string)
                    dampak_list = warning.get('dampak', [])
                    if isinstance(dampak_list, list) and dampak_list:
                        dampak_text = dampak_list[0] if dampak_list else 'Tidak tersedia'
                    else:
                        dampak_text = str(dampak_list)
                    
                    rekomendasi = warning.get('rekomendasi', 'Tidak tersedia')

                    # Isi data ke cell
                    ws.cell(row=current_row, column=1).value = i
                    ws.cell(row=current_row, column=2).value = jenis
                    ws.cell(row=current_row, column=3).value = level.capitalize()
                    ws.cell(row=current_row, column=4).value = warning.get('jumlah', 0)
                    ws.cell(row=current_row, column=5).value = f"{warning.get('persentase', 0):.1f}%"
                    ws.cell(row=current_row, column=6).value = dampak_text[:100]
                    ws.cell(row=current_row, column=7).value = rekomendasi[:150]
                    
                    # Warna cell level
                    ws.cell(row=current_row, column=3).fill = level_colors.get(
                        level, 
                        PatternFill(start_color="FFFFFF", fill_type="solid")
                    )
                    
                    # Border dan alignment
                    for col in range(1, 8):
                        ws.cell(row=current_row, column=col).border = border
                        if col in [6, 7]:  # Kolom dengan teks panjang
                            ws.cell(row=current_row, column=col).alignment = Alignment(
                                wrap_text=True, 
                                vertical='top',
                                horizontal='left'
                            )
                        else:
                            ws.cell(row=current_row, column=col).alignment = center_alignment
                    
                    current_row += 1

                # Set column widths
                ws.column_dimensions['A'].width = 8    # No
                ws.column_dimensions['B'].width = 20   # Jenis
                ws.column_dimensions['C'].width = 15   # Level
                ws.column_dimensions['D'].width = 12   # Jumlah
                ws.column_dimensions['E'].width = 15   # Persentase
                ws.column_dimensions['F'].width = 40   # Dampak
                ws.column_dimensions['G'].width = 50   # Rekomendasi

                current_row += 2

        # ===================== RANKING WILAYAH =====================
        ranking = data.get('ranking_wilayah', {})
        if ranking.get('ranking_terkotor'):
            ws[f'A{current_row}'] = "RANKING WILAYAH - TERKOTOR (TOP 5)"
            ws[f'A{current_row}'].font = Font(bold=True, size=12, color="9B59B6")
            current_row += 2

            headers = ['Peringkat', 'Wilayah', 'Total Laporan', 'Selesai', 'Pending', 'Skor Kebersihan', 'Kategori']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = PatternFill(start_color="9B59B6", fill_type="solid")
                cell.alignment = center_alignment
                cell.border = border
            current_row += 1

            for wilayah in ranking['ranking_terkotor'][:5]:
                ws.cell(row=current_row, column=1).value = wilayah.get('peringkat', 0)
                ws.cell(row=current_row, column=2).value = wilayah.get('wilayah', '')[:40]
                ws.cell(row=current_row, column=3).value = wilayah.get('total_laporan', 0)
                ws.cell(row=current_row, column=4).value = wilayah.get('laporan_selesai', 0)
                ws.cell(row=current_row, column=5).value = wilayah.get('laporan_pending', 0)
                ws.cell(row=current_row, column=6).value = wilayah.get('skor_kebersihan', 0)
                ws.cell(row=current_row, column=7).value = wilayah.get('kategori', '')
                
                # Warna berdasarkan kategori
                kategori_fill = {
                    'Sangat Kotor': PatternFill(start_color="FF0000", fill_type="solid"),
                    'Kotor': PatternFill(start_color="FF6600", fill_type="solid"),
                    'Cukup Bersih': PatternFill(start_color="FFCC00", fill_type="solid"),
                    'Bersih': PatternFill(start_color="99FF99", fill_type="solid"),
                    'Sangat Bersih': PatternFill(start_color="00CC00", fill_type="solid")
                }
                ws.cell(row=current_row, column=7).fill = kategori_fill.get(
                    wilayah.get('kategori', ''), 
                    PatternFill(start_color="FFFFFF", fill_type="solid")
                )
                
                # Border
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).border = border
                    ws.cell(row=current_row, column=col).alignment = center_alignment
                
                current_row += 1

            current_row += 2

        # ===================== REKOMENDASI =====================
        rekom = data.get('rekomendasi', [])
        if rekom:
            ws[f'A{current_row}'] = "REKOMENDASI TINDAKAN"
            ws[f'A{current_row}'].font = Font(bold=True, size=12, color="F39C12")
            current_row += 2

            headers = ['No', 'Prioritas', 'Kategori', 'Rekomendasi', 'Alasan', 'Sumber Data']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = PatternFill(start_color="F39C12", fill_type="solid")
                cell.alignment = center_alignment
                cell.border = border
            current_row += 1

            # Warna prioritas
            prioritas_fill = {
                'SANGAT_TINGGI': PatternFill(start_color="FF0000", fill_type="solid"),
                'TINGGI': PatternFill(start_color="FF9900", fill_type="solid"),
                'SEDANG': PatternFill(start_color="FFFF00", fill_type="solid"),
                'RENDAH': PatternFill(start_color="00FF00", fill_type="solid")
            }

            for i, rec in enumerate(rekom, start=1):
                ws.cell(row=current_row, column=1).value = i
                prio = rec.get('prioritas', 'SEDANG').upper()
                ws.cell(row=current_row, column=2).value = prio
                ws.cell(row=current_row, column=2).fill = prioritas_fill.get(prio, prioritas_fill['SEDANG'])
                ws.cell(row=current_row, column=3).value = rec.get('kategori', '')
                ws.cell(row=current_row, column=4).value = rec.get('rekomendasi', '')
                ws.cell(row=current_row, column=5).value = rec.get('alasan', '')
                ws.cell(row=current_row, column=6).value = rec.get('sumber_data', '')

                # Wrap text untuk kolom dengan teks panjang
                ws.cell(row=current_row, column=4).alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                ws.cell(row=current_row, column=5).alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                ws.cell(row=current_row, column=6).alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                
                # Border untuk semua cell
                for col in range(1, 7):
                    ws.cell(row=current_row, column=col).border = border
                
                current_row += 1

            # Set column widths
            ws.column_dimensions['A'].width = 8    # No
            ws.column_dimensions['B'].width = 15   # Prioritas
            ws.column_dimensions['C'].width = 20   # Kategori
            ws.column_dimensions['D'].width = 60   # Rekomendasi
            ws.column_dimensions['E'].width = 40   # Alasan
            ws.column_dimensions['F'].width = 20   # Sumber Data

            current_row += 2

        # ===================== INFO PERIODE =====================
        ws[f'A{current_row}'] = "INFORMASI LAPORAN"
        ws[f'A{current_row}'].font = Font(bold=True, size=11, color="7F8C8D")
        current_row += 1

        info_row = current_row
        
        # Periode analisis
        if 'period' in data:
            period = data['period']
            if isinstance(period, dict) and 'label' in period:
                ws[f'A{info_row}'] = f"Periode Analisis: {period['label']}"
            else:
                ws[f'A{info_row}'] = f"Periode Analisis: {period}"
        
        # Tanggal generate
        if 'tanggal_generate' in data:
            info_row += 1
            try:
                if isinstance(data['tanggal_generate'], str):
                    tanggal_str = data['tanggal_generate']
                else:
                    tanggal_str = data['tanggal_generate'].strftime('%d %B %Y, %H:%M')
                ws[f'A{info_row}'] = f"Tanggal Generate: {tanggal_str}"
            except:
                ws[f'A{info_row}'] = f"Tanggal Generate: {str(data['tanggal_generate'])}"
        
        # Format info rows
        for row in range(current_row, info_row + 1):
            ws[f'A{row}'].font = Font(size=10, color="7F8C8D")
            ws[f'A{row}'].alignment = Alignment(horizontal='left')

        current_row = info_row + 2

        return current_row

from datetime import datetime, timedelta
from collections import Counter, defaultdict
from django.utils import timezone
from django.core.cache import cache
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied

class DampakLingkunganReportView(ReportViewSet):
    """View untuk laporan dampak lingkungan dari laporan sampah"""
    
    def post(self, request):
        try:
            self.check_admin_permission(request)

            # Validasi input
            start_date = request.data.get('start_date')
            end_date = request.data.get('end_date')

            if not start_date or not end_date:
                return Response({'error': 'start_date dan end_date diperlukan'}, status=400)

            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Format tanggal tidak valid. Gunakan YYYY-MM-DD'}, status=400)

            # Query data laporan sampah dari database - HAPUS select_related
            laporan_qs = LaporanSampah.objects.filter(
                tanggal_lapor__range=[start_date, end_date]
            )

            # Analisis dasar
            total_laporan = laporan_qs.count()
            laporan_selesai = laporan_qs.filter(status='selesai').count()

            # Klasifikasi jenis sampah berdasarkan data asli
            klasifikasi_sampah = self.klasifikasi_jenis_sampah_berdasarkan_data(
                laporan_qs)

            # Analisis wilayah administrasi (kelurahan/desa/kecamatan)
            analisis_wilayah = self.analisis_wilayah_administrasi(laporan_qs)

            # 🆕 Analisis dampak lingkungan berdasarkan JENIS SAMPAH
            dampak_lingkungan = self.analisis_dampak_lingkungan(
                laporan_qs, klasifikasi_sampah)

            # Tren waktu
            tren_waktu = self.analisis_tren_waktu(
                laporan_qs, start_date, end_date)

            # Efektivitas penanganan
            efektivitas = self.hitung_efektivitas_penanganan(laporan_qs)

            # Ranking wilayah terbersih & terkotor
            ranking_wilayah = self.ranking_wilayah_bersih_kotor(
                analisis_wilayah)

            # Prepare response data
            data = {
                'period': self.format_period(start_date, end_date),
                'total_laporan': total_laporan,
                'laporan_selesai': laporan_selesai,
                'tingkat_penyelesaian': (
                    (laporan_selesai / total_laporan *
                     100) if total_laporan > 0 else 0
                ),
                'klasifikasi_sampah': klasifikasi_sampah,
                'analisis_wilayah': analisis_wilayah,
                'dampak_lingkungan': dampak_lingkungan,  # 🆕 Fitur baru
                'ranking_wilayah': ranking_wilayah,
                'tren_waktu': tren_waktu,
                'efektivitas_penanganan': efektivitas,
                'rekomendasi': self.generate_rekomendasi(
                    klasifikasi_sampah, analisis_wilayah, ranking_wilayah, efektivitas, dampak_lingkungan
                ),
                'tanggal_generate': timezone.now()
            }

            return Response(data)

        except PermissionDenied as e:
            return Response({'error': str(e)}, status=403)
        except Exception as e:
            return Response({'error': f'Internal server error: {str(e)}'}, status=500)

    def klasifikasi_jenis_sampah_berdasarkan_data(self, queryset):
        """
        Klasifikasi jenis sampah berdasarkan data asli dari database
        """
        counter = Counter()
        
        for laporan in queryset:
            # Cek berbagai kemungkinan field untuk jenis sampah
            jenis = self.identifikasi_jenis_sampah(laporan)
            counter[jenis] += 1
        
        total = sum(counter.values()) or 1
        
        hasil = []
        for jenis, jumlah in counter.items():
            hasil.append({
                'jenis': jenis,
                'jumlah': jumlah,
                'persentase': round((jumlah / total) * 100, 1),
                'data_valid': jumlah
            })
        
        return {
            'detail_klasifikasi': hasil,
            'total_data_terstruktur': total,
            'persentase_data_terstruktur': 100
        }

    def identifikasi_jenis_sampah(self, laporan):
        """
        Identifikasi jenis sampah dari laporan
        """
        # Cek jika ada field 'jenis' langsung di model LaporanSampah
        if hasattr(laporan, 'jenis') and laporan.jenis:
            return laporan.jenis.lower()
        
        # Fallback ke deteksi dari deskripsi
        deskripsi = (laporan.deskripsi or "").lower()
        return self.deteksi_jenis_sampah_dari_deskripsi(deskripsi)

    def analisis_dampak_lingkungan(self, queryset, klasifikasi_sampah):
        """
        Analisis dampak lingkungan berdasarkan JENIS SAMPAH yang berbahaya
        """
        dampak_data = {
            'total_analisis': queryset.count(),
            'jenis_berbahaya': {},
            'peringatan': [],
            'analisis_detail': []
        }
        
        # Ambil data klasifikasi detail
        detail_klasifikasi = klasifikasi_sampah.get('detail_klasifikasi', [])
        
        # Analisis dampak per jenis sampah
        for item in detail_klasifikasi:
            jenis = item.get('jenis', '')
            jumlah = item.get('jumlah', 0)
            persentase = item.get('persentase', 0)
            
            # Analisis dampak lingkungan berdasarkan jenis sampah
            dampak_info = self.analisis_dampak_per_jenis(jenis, jumlah, persentase, queryset.count())
            
            if dampak_info:
                dampak_data['analisis_detail'].append(dampak_info)
                
                # Jika ada peringatan, tambahkan ke list peringatan
                if dampak_info.get('peringatan_level'):
                    dampak_data['peringatan'].append({
                        'level': dampak_info['peringatan_level'],
                        'jenis': jenis,
                        'jumlah': jumlah,
                        'persentase': persentase,
                        'dampak': dampak_info['dampak_potensial'],
                        'rekomendasi': dampak_info['rekomendasi']
                    })
                
                # Simpan jenis berbahaya
                if dampak_info.get('tingkat_bahaya') in ['sangat_tinggi', 'tinggi', 'sedang']:
                    dampak_data['jenis_berbahaya'][jenis] = {
                        'jumlah': jumlah,
                        'persentase': persentase,
                        'tingkat_bahaya': dampak_info['tingkat_bahaya'],
                        'dampak': dampak_info['dampak_potensial']
                    }
        
        # Analisis lokasi dengan sampah berbahaya
        lokasi_berbahaya = self.analisis_lokasi_berbahaya(queryset)
        if lokasi_berbahaya:
            dampak_data['lokasi_berbahaya'] = lokasi_berbahaya
        
        # Analisis waktu penanganan sampah berbahaya
        waktu_penanganan = self.analisis_waktu_penanganan_berbahaya(queryset)
        if waktu_penanganan:
            dampak_data['waktu_penanganan_berbahaya'] = waktu_penanganan
        
        # Ringkasan
        dampak_data['ringkasan'] = {
            'total_jenis_berbahaya': len(dampak_data['jenis_berbahaya']),
            'total_peringatan': len(dampak_data['peringatan']),
            'tingkat_risiko': self.hitung_tingkat_risiko(dampak_data)
        }
        
        return dampak_data

    def analisis_dampak_per_jenis(self, jenis, jumlah, persentase, total_laporan):
        """
        Analisis dampak lingkungan per jenis sampah
        """
        # Map untuk jenis sampah yang umum
        dampak_map = {
            'b3': {
                'tingkat_bahaya': 'sangat_tinggi',
                'dampak_potensial': [
                    'Pencemaran tanah dan air',
                    'Bahaya kesehatan bagi warga',
                    'Kebakaran atau ledakan',
                    'Racun bagi ekosistem'
                ],
                'rekomendasi': 'Penanganan khusus oleh tim berwenang, pemisahan limbah, edukasi masyarakat',
                'ambang_peringatan': 1  # 1 laporan sudah perlu peringatan
            },
            'plastik': {
                'tingkat_bahaya': 'sedang',
                'dampak_potensial': [
                    'Pencemaran tanah jangka panjang',
                    'Menyumbat saluran air',
                    'Bahaya bagi hewan',
                    'Mikroplastik di lingkungan'
                ],
                'rekomendasi': 'Program daur ulang, pengurangan penggunaan plastik, bank sampah',
                'ambang_peringatan': 30  # >30% dari total laporan
            },
            'organik': {
                'tingkat_bahaya': 'rendah',
                'dampak_potensial': [
                    'Bau tidak sedap',
                    'Menyebabkan penyakit',
                    'Mengundang hama',
                    'Pencemaran udara'
                ],
                'rekomendasi': 'Pengomposan, pengangkutan rutin, TPS terpadu',
                'ambang_peringatan': 40  # >40% dari total laporan
            },
            'logam': {
                'tingkat_bahaya': 'sedang',
                'dampak_potensial': [
                    'Luka fisik',
                    'Karat mencemari tanah',
                    'Bahaya bagi anak-anak'
                ],
                'rekomendasi': 'Daur ulang, pengumpulan terpisah',
                'ambang_peringatan': 20  # >20% dari total laporan
            },
            'kaca': {
                'tingkat_bahaya': 'sedang',
                'dampak_potensial': [
                    'Luka fisik',
                    'Bahaya bagi hewan',
                    'Sulit terurai'
                ],
                'rekomendasi': 'Pengumpulan terpisah, daur ulang',
                'ambang_peringatan': 15  # >15% dari total laporan
            },
            'campuran': {
                'tingkat_bahaya': 'rendah',
                'dampak_potensial': [
                    'Kesulitan daur ulang',
                    'Peningkatan volume TPA',
                    'Potensi pencemaran campuran'
                ],
                'rekomendasi': 'Program pemilahan sampah, edukasi masyarakat',
                'ambang_peringatan': 50  # >50% dari total laporan
            },
            'kertas': {
                'tingkat_bahaya': 'rendah',
                'dampak_potensial': [
                    'Mudah terbakar',
                    'Menyumbat saluran air',
                    'Rawan penyakit saat basah'
                ],
                'rekomendasi': 'Daur ulang, bank sampah',
                'ambang_peringatan': 25  # >25% dari total laporan
            }
        }
        
        # Jika jenis tidak dikenali, gunakan default
        if jenis not in dampak_map:
            # Default untuk jenis tidak dikenal
            info = {
                'tingkat_bahaya': 'rendah',
                'dampak_potensial': ['Dampak lingkungan perlu investigasi lebih lanjut'],
                'rekomendasi': 'Identifikasi jenis sampah yang lebih spesifik',
                'ambang_peringatan': 50
            }
        else:
            info = dampak_map[jenis]
        
        # Tentukan level peringatan
        peringatan_level = None
        if jenis == 'b3' and jumlah > 0:
            peringatan_level = 'sangat_tinggi'
        elif persentase > info['ambang_peringatan']:
            if info['tingkat_bahaya'] == 'sangat_tinggi':
                peringatan_level = 'sangat_tinggi'
            elif info['tingkat_bahaya'] == 'tinggi':
                peringatan_level = 'tinggi'
            elif info['tingkat_bahaya'] == 'sedang':
                peringatan_level = 'sedang'
            elif info['tingkat_bahaya'] == 'rendah' and persentase > 50:
                peringatan_level = 'rendah'
        
        return {
            'jenis': jenis,
            'jumlah': jumlah,
            'persentase': persentase,
            'tingkat_bahaya': info['tingkat_bahaya'],
            'dampak_potensial': info['dampak_potensial'],
            'rekomendasi': info['rekomendasi'],
            'ambang_peringatan': info['ambang_peringatan'],
            'peringatan_level': peringatan_level,
            'status': 'aman' if not peringatan_level else 'perhatian'
        }

    def analisis_lokasi_berbahaya(self, queryset):
        """
        Analisis lokasi dengan konsentrasi sampah berbahaya tinggi
        """
        lokasi_berbahaya = []
        
        # Kelompokkan per wilayah berdasarkan koordinat (grid)
        from collections import defaultdict
        lokasi_counter = defaultdict(lambda: {
            'b3': 0,
            'plastik': 0,
            'logam': 0,
            'kaca': 0,
            'total': 0,
            'alamat_samples': [],
            'koordinat': None
        })
        
        for laporan in queryset:
            jenis = self.identifikasi_jenis_sampah(laporan)
            
            if jenis in ['b3', 'plastik', 'logam', 'kaca']:
                try:
                    lat = round(float(laporan.latitude), 3) if laporan.latitude else None
                    lon = round(float(laporan.longitude), 3) if laporan.longitude else None
                    
                    if lat and lon:
                        # Buat key berdasarkan grid koordinat (presisi 3 desimal ≈ 111m)
                        grid_key = f"{lat:.3f},{lon:.3f}"
                        
                        lokasi_counter[grid_key][jenis] += 1
                        lokasi_counter[grid_key]['total'] += 1
                        
                        # Simpan alamat sample
                        if len(lokasi_counter[grid_key]['alamat_samples']) < 3:
                            lokasi_counter[grid_key]['alamat_samples'].append(laporan.alamat[:50] if laporan.alamat else "Alamat tidak tersedia")
                        
                        # Simpan koordinat
                        if not lokasi_counter[grid_key]['koordinat']:
                            lokasi_counter[grid_key]['koordinat'] = {
                                'lat': lat,
                                'lon': lon
                            }
                except (ValueError, TypeError):
                    continue
        
        # Filter lokasi dengan sampah berbahaya signifikan
        for grid_key, data in lokasi_counter.items():
            if data['b3'] > 0 or data['plastik'] > 5 or data['total'] > 10:
                lokasi_berbahaya.append({
                    'lokasi': f"Koordinat: {data['koordinat']['lat']:.4f}, {data['koordinat']['lon']:.4f}" if data['koordinat'] else 'Lokasi tidak diketahui',
                    'b3': data['b3'],
                    'plastik': data['plastik'],
                    'logam': data['logam'],
                    'kaca': data['kaca'],
                    'total_berbahaya': data['total'],
                    'alamat_samples': data['alamat_samples'],
                    'koordinat': data['koordinat'],
                    'tingkat_risiko': 'tinggi' if data['b3'] > 0 else 'sedang'
                })
        
        # Urutkan berdasarkan tingkat risiko
        lokasi_berbahaya.sort(key=lambda x: (1 if x['tingkat_risiko'] == 'tinggi' else 0, x['total_berbahaya']), reverse=True)
        
        return lokasi_berbahaya[:10]

    def analisis_waktu_penanganan_berbahaya(self, queryset):
        """
        Analisis waktu penanganan sampah berbahaya
        """
        # Cek jika ada field tanggal_selesai
        if not hasattr(queryset.first(), 'tanggal_selesai'):
            return None
        
        laporan_berbahaya = []
        for laporan in queryset:
            jenis = self.identifikasi_jenis_sampah(laporan)
            if jenis in ['b3', 'plastik', 'logam', 'kaca']:
                if hasattr(laporan, 'tanggal_selesai') and laporan.tanggal_selesai and laporan.tanggal_lapor:
                    try:
                        waktu_penanganan = (laporan.tanggal_selesai - laporan.tanggal_lapor).days
                        laporan_berbahaya.append({
                            'jenis': jenis,
                            'waktu_hari': waktu_penanganan,
                            'status': laporan.status
                        })
                    except (TypeError, AttributeError):
                        continue
        
        if not laporan_berbahaya:
            return None
        
        # Hitung statistik
        total = len(laporan_berbahaya)
        selesai = sum(1 for x in laporan_berbahaya if x['status'] == 'selesai')
        if selesai > 0:
            rata_waktu = sum(x['waktu_hari'] for x in laporan_berbahaya if x['status'] == 'selesai') / selesai
        else:
            rata_waktu = 0
        
        return {
            'total_laporan_berbahaya': total,
            'selesai': selesai,
            'pending': total - selesai,
            'rata_waktu_penanganan_hari': round(rata_waktu, 1),
            'tingkat_penanganan': round((selesai / total) * 100, 1) if total > 0 else 0
        }

    def hitung_tingkat_risiko(self, dampak_data):
        """
        Hitung tingkat risiko keseluruhan
        """
        peringatan_sangat_tinggi = sum(1 for p in dampak_data.get('peringatan', []) if p['level'] == 'sangat_tinggi')
        peringatan_tinggi = sum(1 for p in dampak_data.get('peringatan', []) if p['level'] == 'tinggi')
        peringatan_sedang = sum(1 for p in dampak_data.get('peringatan', []) if p['level'] == 'sedang')
        
        if peringatan_sangat_tinggi > 0:
            return 'sangat_tinggi'
        elif peringatan_tinggi > 0:
            return 'tinggi'
        elif peringatan_sedang > 2:
            return 'sedang'
        elif dampak_data.get('jenis_berbahaya'):
            return 'rendah'
        else:
            return 'aman'
        
    def deteksi_jenis_sampah_dari_deskripsi(self, deskripsi):
        """Deteksi cepat jenis sampah dari deskripsi pendek"""
        keywords = {
            # ==================== B3 (BAHAN BERBAHAYA & BERACUN) ====================
            'b3': [
                # B3 Umum
                'b3', 'berbahaya', 'beracun', 'toxic', 'hazardous', 'limbah b3',
                'bahan berbahaya', 'bahan beracun', 'limbah berbahaya', 
                'limbah beracun', 'zat berbahaya', 'zat beracun',
                
                # B3 Medis/Kesehatan
                'jarum suntik', 'suntik', 'infus bekas', 'medis', 'kesehatan',
                'perban bekas', 'alat medis', 'sarana kesehatan', 'klinik',
                'rumah sakit', 'laboratorium', 'spuit', 'needle', 'syringe',
                'iv set', 'kateter', 'sarung tangan medis', 'masker medis',
                'pembalut medis', 'plester medis', 'alkohol medis', 'betadine',
                'obat suntik', 'vaksin', 'sampah medis', 'limbah medis',
                'limbah infeksius', 'limbah patologi', 'limbah farmasi',
                
                # B3 Elektronik/E-Waste
                'elektronik', 'e-waste', 'lampu neon', 'lampu hemat energi',
                'lampu led rusak', 'lampu tl', 'lampu fluorescent',
                'komputer rusak', 'laptop rusak', 'tv rusak', 'kulkas rusak',
                'ac rusak', 'ponsel rusak', 'hp rusak', 'tablet rusak',
                'baterai', 'aki', 'accu', 'accumulator', 'baterai mobil',
                'baterai motor', 'baterai lithium', 'baterai isi ulang',
                'crt', 'monitor rusak', 'pc rusak', 'printer rusak',
                'scanner rusak', 'kabel listrik', 'transformator', 'trafo',
                'power supply', 'charger rusak', 'adaptor rusak',
                'microwave rusak', 'oven rusak', 'blender rusak',
                
                # B3 Oli & Pelumas
                'oli bekas', 'pelumas', 'minyak mesin', 'gemuk', 'lumas',
                'oli motor', 'oli mobil', 'oli industri', 'oli hydraulic',
                'oli transmisi', 'oli gardan', 'oli rem', 'oli sampah',
                'minyak bekas', 'gemuk bekas', 'grease', 'oli sintetis',
                
                # B3 Cat & Pelarut
                'cat bekas', 'thinner', 'vernis', 'pelarut', 'solvent',
                'aerosol', 'semprot', 'spray can', 'kaleng cat', 'kaleng semprot',
                'cat minyak', 'cat tembok', 'cat kayu', 'cat besi',
                'lak', 'pengencer', 'pembersih cat', 'penghapus cat',
                'acetone', 'tiner', 'spiritus', 'alkohol teknis',
                
                # B3 Logam Berat
                'merkuri', 'timbal', 'kadmium', 'logam berat', 'air raksa',
                'raksa', 'arsen', 'kromium', 'nikel', 'seng', 'tembaga beracun',
                'aluminium beracun', 'besi beracun', 'solder timah',
                'timah hitam', 'timah putih', 'sianida', 'sianida',
                
                # B3 Gas & Tabung
                'tabung gas', 'aerosol', 'freon', 'refrigerant', 'kaleng semprot',
                'gas', 'elpiji bekas', 'tabung elpiji', 'tabung oksigen',
                'tabung nitrogen', 'tabung co2', 'tabung las', 'tabung ac',
                'korek gas', 'lighters', 'pemantik', 'butane', 'propane',
                
                # B3 Asbes & Konstruksi
                'asbes', 'atap asbes', 'serat asbes', 'asbestos', 'gypsum',
                'bahan bangunan berbahaya', 'cat timbal', 'cat mengandung timbal',
                
                # B3 Farmasi
                'obat kadaluarsa', 'obat rusak', 'farmasi', 'obat bekas',
                'vaksin kadaluarsa', 'antibiotik kadaluarsa', 'sirup kadaluarsa',
                'tablet kadaluarsa', 'kapsul kadaluarsa', 'obat cair kadaluarsa',
                'suplemen kadaluarsa', 'vitamin kadaluarsa', 'obat resep',
                'obat keras', 'psikotropika', 'narkotika', 'obat terlarang',
                
                # B3 Pertanian
                'pestisida', 'herbisida', 'insektisida', 'fungisida',
                'racun tikus', 'racun serangga', 'urea kadaluarsa', 'pupuk kadaluarsa',
                'pupuk kimia', 'zat perangsang', 'hormon tanaman', 'antibiotik hewan',
                'vaksin hewan', 'obat hewan kadaluarsa', 'desinfektan',
                
                # B3 Industri
                'limbah pabrik', 'limbah industri', 'slag', 'abu industri',
                'sludge', 'limbah cair berbahaya', 'limbah padat berbahaya',
                'limbah gas berbahaya', 'cairan kimia', 'bahan kimia industri',
                'acid', 'asam', 'basa', 'alkali', 'detergen industri',
                'pemutih industri', 'pelarut industri', 'catalyst',
                'resin', 'polimer berbahaya', 'plastik pvc',
                
                # B3 Rumah Tangga
                'pembalut wanita', 'pembalut bekas', 'popok bekas', 'diapers',
                'tissue bekas darah', 'tissue medis', 'kapas medis',
                'pembersih lantai', 'pembersih toilet', 'pemutih pakaian',
                'pengharum ruangan', 'insektisida rumah', 'anti nyamuk',
                'obat nyamuk', 'repellent', 'racun kecoa', 'racun semut',
            ],
            
            # ==================== PLASTIK ====================
            'plastik': [
                # Plastik Umum
                'plastik', 'botol plastik', 'gelas plastik', 'kresek', 'kantong plastik',
                'kemasan plastik', 'bungkus plastik', 'plastik kemasan',
                'plastik pembungkus', 'plastik belanja', 'tas plastik',
                'plastik tipis', 'plastik tebal', 'plastik transparan',
                'plastik berwarna', 'plastik bening', 'plastik putih',
                
                # Jenis Plastik Spesifik
                'pet', 'hdpe', 'pvc', 'ldpe', 'pp', 'ps', 'other',
                'polyethylene', 'polypropylene', 'polystyrene', 'polyvinyl',
                'nylon', 'polyester', 'acrylic', 'polycarbonate',
                
                # Botol & Wadah Plastik
                'botol air mineral', 'botol minuman', 'botol soda',
                'botol jus', 'botol sirup', 'botol kecap', 'botol saus',
                'botol sampo', 'botol sabun', 'botol deterjen',
                'botol minyak', 'botol obat', 'botol vitamin',
                'galon', 'jerigen', 'drum plastik', 'ember plastik',
                'bak plastik', 'wadah plastik', 'kontainer plastik',
                'tupperware', 'tempat makan plastik', 'kotak plastik',
                
                # Kemasan Makanan Plastik
                'styrofoam', 'gabus plastik', 'bungkus makanan',
                'plastik wrap', 'cling wrap', 'plastik pembungkus makanan',
                'kemasan snack', 'bungkus permen', 'bungkus coklat',
                'bungkus mi instan', 'bungkus kopi', 'bungkus teh',
                'sachet', 'bungkus kecil', 'pouch',
                
                # Peralatan Plastik
                'sedotan', 'straw', 'sendok plastik', 'garpu plastik',
                'pisau plastik', 'piring plastik', 'mangkuk plastik',
                'gelas plastik', 'cup plastik', 'tutup plastik',
                'stirrer', 'pengaduk plastik', 'tusuk gigi plastik',
                'sumpit plastik', 'tutup gelas', 'tutup botol plastik',
                
                # Plastik Konstruksi & Rumah Tangga
                'pipa plastik', 'paralon', 'talang plastik', 'pipa pvc',
                'pipa hdpe', 'plastik cor', 'terpal plastik', 'plastik mulsa',
                'plastik tanaman', 'pot plastik', 'polybag',
                'plastik sampah', 'kantong sampah', 'trash bag',
                'plastik hitam', 'plastik biru', 'plastik merah',
                
                # Mainan & Perlengkapan Plastik
                'mainan plastik', 'boneka plastik', 'lego', 'blok plastik',
                'ember mainan', 'bak mandi plastik', 'kursi plastik',
                'meja plastik', 'rak plastik', 'lemari plastik',
                'box plastik', 'container plastik', 'organizer plastik',
            ],
            
            # ==================== ORGANIK ====================
            'organik': [
                # Sisa Makanan
                'sisa makanan', 'makanan basi', 'makanan kadaluarsa',
                'nasi basi', 'nasi sisa', 'roti basi', 'kue basi',
                'sayur basi', 'buah busuk', 'daging busuk', 'ikan busuk',
                'ayam busuk', 'telur busuk', 'susu basi', 'keju busuk',
                'makanan berjamur', 'makanan fermentasi', 'makanan terbuang',
                
                # Buah & Sayuran
                'buah', 'sayur', 'daun', 'tumbuhan', 'tanaman',
                'kulit buah', 'kulit sayur', 'biji buah', 'biji sayur',
                'batang', 'ranting', 'dahan', 'ranting pohon',
                'daun kering', 'daun basah', 'daun gugur',
                'pepaya', 'pisang', 'apel', 'jeruk', 'mangga',
                'semangka', 'melon', 'anggur', 'strawberry',
                'bayam', 'kangkung', 'sawi', 'wortel', 'kentang',
                'tomat', 'cabe', 'bawang', 'jahe', 'kunyit',
                
                # Dedaunan & Tanaman
                'rumput', 'rumput potong', 'rumput liar',
                'ranting kecil', 'ranting besar', 'dahan pohon',
                'bambu', 'daun bambu', 'batang pisang', 'pelepah',
                'batang jagung', 'batang singkong', 'batang ubi',
                'akar', 'umbi', 'rimpang',
                
                # Bahan Organik Lainnya
                'telur', 'cangkang telur', 'kulit telur',
                'tulang', 'tulang ayam', 'tulang ikan', 'tulang sapi',
                'cangkang', 'cangkang kerang', 'cangkang kepiting',
                'sisik ikan', 'kepala ikan', 'insang ikan',
                'bulu', 'bulu ayam', 'bulu hewan', 'rambut',
                'kotoran hewan', 'kotoran sapi', 'kotoran kambing',
                'kotoran ayam', 'pupuk kandang', 'kompos',
                
                # Organik Dapur
                'ampas kopi', 'ampas teh', 'serbuk kayu',
                'bumbu dapur', 'rempah', 'merica', 'ketumbar',
                'bawang merah', 'bawang putih', 'bawang bombay',
                'daun bawang', 'seledri', 'kemangi', 'daun salam',
                'serai', 'lengkuas', 'kencur', 'temulawak',
                
                # Organik Kebun
                'dedaunan', 'ranting pohon', 'batang kecil',
                'bunga', 'bunga layu', 'bunga gugur',
                'tanaman mati', 'tanaman layu', 'tanaman sakit',
                'potongan rumput', 'clipping', 'trimming',
                'gulma', 'tanaman liar', 'tumbuhan liar',
            ],
            
            # ==================== KERTAS ====================
            'kertas': [
                # Kertas Umum
                'kertas', 'kertas bekas', 'kertas koran', 'koran',
                'majalah', 'tabloid', 'buku', 'buku bekas',
                'novel', 'komik', 'majalah bekas', 'kertas hvs',
                'kertas folio', 'kertas a4', 'kertas f4',
                'kertas buram', 'kertas sampul', 'kertas kado',
                'kertas warna', 'kertas karton', 'karton',
                
                # Kertas Kemasan
                'kardus', 'kardus bekas', 'box kardus',
                'dus', 'dus kardus', 'kotak kardus',
                'kemasan kardus', 'pembungkus kardus',
                'karton box', 'karton dus', 'paper bag',
                'tas kertas', 'kantong kertas', 'bungkus kertas',
                'kertas minyak', 'kertas roti', 'kertas nasi',
                'kertas pembungkus', 'kertas kemasan',
                
                # Kertas Kantor
                'dokumen', 'arsip', 'file', 'laporan',
                'surat', 'nota', 'faktur', 'invoice',
                'kwitansi', 'struk', 'tiket', 'karcis',
                'formulir', 'lembar kerja', 'worksheet',
                'print out', 'hasil print', 'fotokopi',
                'printan', 'print-an', 'cetakan',
                
                # Kertas Rumah Tangga
                'tissue', 'tisu', 'kertas tissue', 'kertas tisu',
                'tissue toilet', 'tissue wajah', 'tissue dapur',
                'serviet', 'napkin', 'handuk kertas',
                'kertas rokok', 'bungkus rokok', 'kemasan rokok',
                'bungkus gula', 'bungkus tepung', 'bungkus garam',
                
                # Kertas Spesial
                'kertas foto', 'foto', 'print foto',
                'poster', 'brosur', 'leaflet', 'pamflet',
                'flyer', 'spanduk kertas', 'banner kertas',
                'kalender', 'kalender bekas', 'agenda',
                'notes', 'buku catatan', 'notebook',
                
                # Kertas Industri
                'kertas kraft', 'kertas packing',
                'kertas semen', 'kertas gipsum',
                'kertas duplex', 'kertas ivory',
                'kertas art paper', 'kertas art carton',
                'kertas sticker', 'label', 'stiker',
            ],
            
            # ==================== LOGAM ====================
            'logam': [
                # Logam Umum
                'logam', 'besi', 'baja', 'steel', 'stainless',
                'aluminium', 'alumunium', 'alumimum',
                'tembaga', 'copper', 'kuningan', 'brass',
                'perunggu', 'bronze', 'timah', 'tin',
                'seng', 'zinc', 'nikel', 'nickel',
                'krom', 'chromium', 'magnesium',
                
                # Kaleng & Kemasan Logam
                'kaleng', 'kaleng bekas', 'kaleng minuman',
                'kaleng soda', 'kaleng bir', 'kaleng susu',
                'kaleng cat', 'kaleng makanan', 'kaleng sarden',
                'kaleng kornet', 'kaleng susu kental',
                'kemasan kaleng', 'wadah kaleng',
                
                # Peralatan Rumah Tangga Logam
                'panci', 'wajan', 'teflon', 'kuali',
                'sendok logam', 'garpu logam', 'pisau logam',
                'sodet', 'spatula', 'saringan',
                'ember logam', 'bak logam', 'bucket',
                'keranjang logam', 'rak logam',
                'gantungan baju logam', 'hanger',
                
                # Perkakas & Konstruksi
                'paku', 'sekrup', 'baut', 'mur',
                'kawat', 'kawat berduri', 'kawat ayam',
                'kawat bendrat', 'kabel listrik',
                'besi beton', 'besi cor', 'besi hollow',
                'besi siku', 'besi plat', 'plat besi',
                'pipa besi', 'pipa galvanis', 'pipa tembaga',
                'paralon besi', 'talang seng',
                
                # Elektronik & Komponen
                'motor listrik', 'dinamo', 'generator',
                'transformator', 'trafo', 'kumparan',
                'kawat tembaga', 'kabel bekas', 'kabel listrik',
                'kabel telepon', 'kabel coaxial', 'kabel usb',
                'pcb', 'printed circuit board', 'komponen elektronik',
                'chip', 'processor', 'ram', 'harddisk',
                
                # Kendaraan & Mesin
                'velg', 'roda', 'rantai', 'gear',
                'mesin', 'engine', 'blok mesin',
                'knalpot', 'exhaust', 'karburator',
                'radiator', 'alternator', 'starter',
                'body mobil', 'body motor', 'chassis',
                
                # Furniture & Dekorasi
                'ranjang besi', 'tempat tidur besi',
                'kursi besi', 'meja besi', 'lemari besi',
                'pagar besi', 'teralis', 'jeruji',
                'kanopi besi', 'atap seng', 'seng',
                'genteng metal', 'atap metal',
            ],
            
            # ==================== KACA ====================
            'kaca': [
                # Kaca Umum
                'kaca', 'beling', 'pecahan kaca', 'kaca pecah',
                'gelas kaca', 'gelas beling', 'piring kaca',
                'mangkuk kaca', 'cangkir kaca', 'teko kaca',
                'vas kaca', 'botol kaca', 'botol beling',
                
                # Botol Kaca Spesifik
                'botol sirup', 'botol kecap', 'botol saus',
                'botol selai', 'botol madu', 'botol vitamin',
                'botol obat', 'botol parfum', 'botol kosmetik',
                'botol minuman', 'botol bir', 'botol anggur',
                'botol champagne', 'botol spirit',
                'botol susu', 'botol bayi', 'botol dot',
                
                # Peralatan Dapur Kaca
                'toples kaca', 'jar kaca', 'wadah kaca',
                'container kaca', 'mangkuk kaca', 'piring kaca',
                'gelas minum', 'gelas wine', 'gelas cocktail',
                'gelas shot', 'gelas beer', 'gelas juice',
                'teko', 'ceret kaca', 'kendi kaca',
                
                # Kaca Jendela & Bangunan
                'kaca jendela', 'jendela kaca', 'kaca pintu',
                'kaca mobil', 'kaca motor', 'kaca spion',
                'kaca cermin', 'cermin', 'kaca reflektor',
                'kaca film', 'kaca tempered', 'kaca laminasi',
                'kaca patri', 'stained glass',
                
                # Elektronik & Peralatan
                'lampu pijar', 'lampu bohlam', 'lampu halogen',
                'tabung tv', 'crt monitor', 'tabung neon',
                'kaca mikroskop', 'kaca teropong', 'lensa',
                'kacamata', 'spectacles', 'lensa kacamata',
                'kaca pembesar', 'magnifying glass',
                
                # Dekorasi & Seni
                'vas bunga', 'pot kaca', 'aquarium',
                'terarium', 'display case', 'etalse kaca',
                'pigura kaca', 'frame kaca', 'plakat kaca',
                'trophy', 'piala', 'medali dengan kaca',
            ],
            
            # ==================== CAMPURAN ====================
            'campuran': [
                # Umum
                'campuran', 'bermacam', 'beragam', 'beraneka',
                'sampah rumah tangga', 'sampah dapur',
                'sampah kebun', 'sampah taman',
                'sampah kantor', 'sampah sekolah',
                'sampah pasar', 'sampah komersial',
                
                # Rumah Tangga Campuran
                'sampah rumah', 'sampah keluarga',
                'sampah sehari-hari', 'sampah harian',
                'sampah basah kering', 'sampah basah',
                'sampah kering', 'sampah residu',
                
                # Spesifik Lokasi
                'sampah sekolah', 'sampah kampus',
                'sampah kantor', 'sampah perkantoran',
                'sampah pabrik', 'sampah industri',
                'sampah pasar', 'sampah tradisional',
                'sampah mall', 'sampah pusat perbelanjaan',
                'sampah hotel', 'sampah restoran',
                'sampah kafe', 'sampah warung',
                
                # Tak Terklasifikasi
                'sampah tak terpilah', 'sampah tercampur',
                'sampah tidak terpisah', 'sampah gabungan',
                'sampah all in', 'sampah semua jenis',
                'sampah berbagai jenis', 'sampah heterogen',
            ],
            
            # ==================== KARET ====================
            'karet': [
                # Ban
                'ban', 'ban bekas', 'ban mobil', 'ban motor',
                'ban sepeda', 'ban truk', 'ban bus',
                'ban luar', 'ban dalam', 'tube',
                'ban vulkanisir', 'ban recapan',
                
                # Alas Kaki
                'sandal', 'sandal bekas', 'sandal jepit',
                'sepatu', 'sepatu bekas', 'sepatu kets',
                'sepatu olahraga', 'boots', 'sepatu boot',
                'sepatu kulit sintetis', 'sepatu karet',
                
                # Peralatan Karet
                'tali karet', 'karet gelang', 'gelang karet',
                'rubber band', 'elastic band', 'karet pentil',
                'karet penghapus', 'penghapus', 'eraser',
                'karet sandal', 'sol karet', 'sole',
                
                # Karet Industri
                'belt', 'conveyor belt', 'fan belt',
                'timing belt', 'v-belt', 'karet mesin',
                'gasket', 'seal', 'oring', 'o-ring',
                'rubber sheet', 'lembaran karet',
                'karet busa', 'foam rubber',
                
                # Mainan & Perlengkapan
                'balon', 'balloon', 'balon karet',
                'mainan karet', 'rubber toy', 'bola karet',
                'basketball', 'volleyball', 'football',
                'karet stress ball', 'anti stress ball',
                
                # Karet Rumah Tangga
                'sarung tangan karet', 'glove karet',
                'karet pel', 'pel karet', 'squeegee',
                'karet pintu', 'door seal', 'weather strip',
                'karet jendela', 'window seal',
            ],
            
            # ==================== TEKSTIL ====================
            'tekstil': [
                # Pakaian
                'baju', 'pakaian', 'clothing', 'apparel',
                'kaos', 't-shirt', 'kemeja', 'shirt',
                'celana', 'pants', 'jeans', 'jins',
                'rok', 'skirt', 'dress', 'gaun',
                'jaket', 'jacket', 'sweater', 'hoodie',
                'kaus kaki', 'socks', 'stocking',
                'dalaman', 'underwear', 'bra', 'bh',
                'pakaian dalam', 'inner wear',
                
                # Kain & Bahan
                'kain', 'textile', 'fabric', 'cloth',
                'kain perca', 'kain sisa', 'kain bekas',
                'kain potongan', 'scrap fabric',
                'kain katun', 'cotton', 'kain sutra', 'silk',
                'kain wol', 'wool', 'kain linen', 'linen',
                'kain polyester', 'polyester', 'nylon',
                'kain denim', 'denim', 'kain jeans',
                
                # Sprei & Perlengkapan Tidur
                'sprei', 'bedsheet', 'sarung bantal',
                'pillow case', 'sarung guling',
                'selimut', 'blanket', 'bedcover',
                'quilt', 'bedspread', 'kain tempat tidur',
                
                # Handuk & Kain Lap
                'handuk', 'towel', 'handuk mandi',
                'handuk kecil', 'face towel',
                'handuk dapur', 'kitchen towel',
                'kain lap', 'lap', 'rag', 'kain pel',
                'kain bersih', 'cleaning cloth',
                
                # Tas & Aksesoris
                'tas', 'bag', 'tas kain', 'cloth bag',
                'tas belanja', 'shopping bag',
                'tas tangan', 'handbag', 'tas ransel',
                'backpack', 'tas sekolah', 'school bag',
                'topi', 'hat', 'cap', 'kupluk',
                'sarung tangan', 'gloves', 'mittens',
                'syal', 'scarf', 'shawl',
                
                # Gorden & Dekorasi
                'gorden', 'curtain', 'tirai', 'blind',
                'korden', 'window curtain', 'kain penutup',
                'taplak', 'table cloth', 'table runner',
                'karpet', 'carpet', 'rug', 'keset',
                'doormat', 'welcome mat',
            ],
            
            # ==================== KAYU ====================
            'kayu': [
                # Kayu Umum
                'kayu', 'wood', 'kayu bekas', 'kayu sisa',
                'potongan kayu', 'serpihan kayu', 'serbuk kayu',
                'ranting kayu', 'dahan kayu', 'batang kayu',
                
                # Furniture Kayu
                'meja kayu', 'kursi kayu', 'lemari kayu',
                'rak kayu', 'tempat tidur kayu', 'bangku kayu',
                'bufet kayu', 'kitchen set kayu',
                
                # Konstruksi & Bangunan
                'papan', 'balok', 'kasau', 'reng',
                'kayu lapis', 'plywood', 'triplek',
                'multiplek', 'blockboard', 'particle board',
                'mdf', 'hdf', 'hardboard',
                
                # Kemasan Kayu
                'palet', 'pallet', 'dus kayu', 'peti kayu',
                'crate', 'kotak kayu', 'box kayu',
                'kemasan kayu', 'packing kayu',
                
                # Peralatan & Perlengkapan
                'gagang sapu', 'gagang perkakas',
                'tongkat', 'stick', 'batang',
                'pohon tumbang', 'kayu bakar', 'firewood',
                'ranting kering', 'dahan kering',
            ],
            
            # ==================== KERAMIK ====================
            'keramik': [
                'keramik', 'ceramic', 'porselen', 'porcelain',
                'piring keramik', 'mangkuk keramik', 'gelas keramik',
                'guci', 'vas keramik', 'pot keramik',
                'ubin', 'tile', 'lantai keramik', 'wall tile',
                'kloset', 'toilet', 'washtafel', 'sink',
                'bathtub', 'bath tub', 'shower tray',
            ],
            
            # ==================== ELEKTRONIK ====================
            'elektronik': [
                # Umum
                'elektronik', 'e-waste', 'limbah elektronik',
                'barang elektronik rusak', 'perangkat elektronik',
                
                # Komputer & Aksesoris
                'komputer', 'pc', 'laptop', 'notebook',
                'monitor', 'crt', 'lcd', 'led monitor',
                'keyboard', 'mouse', 'printer', 'scanner',
                'harddisk', 'hdd', 'ssd', 'flashdisk',
                'cd', 'dvd', 'bluray', 'optical disc',
                
                # Telekomunikasi
                'ponsel', 'hp', 'smartphone', 'tablet',
                'ipad', 'telepon', 'telepon rumah',
                'modem', 'router', 'switch', 'hub',
                
                # Peralatan Rumah Tangga
                'tv', 'televisi', 'kulkas', 'refrigerator',
                'ac', 'air conditioner', 'kipas angin',
                'blender', 'mixer', 'juicer', 'food processor',
                'microwave', 'oven', 'toaster', 'rice cooker',
                'dispenser', 'water dispenser',
                
                # Audio & Video
                'radio', 'speaker', 'sound system',
                'dvd player', 'bluray player', 'game console',
                'playstation', 'xbox', 'nintendo',
                'kamera', 'camera', 'camcorder', 'video camera',
                
                # Baterai & Power
                'baterai', 'battery', 'aki', 'accu',
                'charger', 'adaptor', 'power supply',
            ],
            
            # ==================== MEDIS ====================
            'medis': [
                'medis', 'kesehatan', 'sarana kesehatan',
                'rumah sakit', 'klinik', 'puskesmas',
                'laboratorium', 'apotek', 'farmasi',
                'obat', 'vaksin', 'antibiotik',
                'perban', 'kasa', 'kapas',
                'jarum', 'suntik', 'infus',
                'sarung tangan', 'masker',
                'pembalut', 'popok',
            ]
        }
        
        deskripsi = deskripsi.lower()
        for jenis, kata_kunci in keywords.items():
            if any(kata in deskripsi for kata in kata_kunci):
                return jenis
        
        return 'tidak_terdeteksi'

    def format_period(self, start_date, end_date):
        return {
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'label': f"{start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}"
        }

    def analisis_tren_waktu(self, queryset, start_date, end_date):
        """
        Analisis tren laporan sampah berdasarkan waktu (harian)
        Output siap untuk chart (line / bar)
        """
        # Inisialisasi range tanggal
        delta = (end_date - start_date).days
        date_range = [
            start_date + timedelta(days=i)
            for i in range(delta + 1)
        ]

        # Siapkan struktur default
        tren = {
            'labels': [d.strftime('%Y-%m-%d') for d in date_range],
            'total_laporan': [],
            'laporan_selesai': [],
            'laporan_pending': []
        }

        # Pre-calc queryset
        qs = queryset.values('tanggal_lapor', 'status')

        for tanggal in date_range:
            harian = [q for q in qs if q['tanggal_lapor'] == tanggal]

            total = len(harian)
            selesai = sum(1 for q in harian if q['status'] == 'selesai')
            pending = total - selesai

            tren['total_laporan'].append(total)
            tren['laporan_selesai'].append(selesai)
            tren['laporan_pending'].append(pending)

        return tren

    def analisis_wilayah_administrasi(self, queryset):
        """
        Analisis wilayah administrasi - VERSI SIMPLIFIED
        """
        wilayah_data = []
        
        for laporan in queryset[:50]:  # Batasi untuk performa
            try:
                lat = float(laporan.latitude) if laporan.latitude else None
                lon = float(laporan.longitude) if laporan.longitude else None
                
                # SIMPLIFIED: Ekstrak informasi dari alamat (jika ada)
                alamat = laporan.alamat or ""
                
                # Coba ekstrak kecamatan/kelurahan dari alamat jika memungkinkan
                # atau gunakan informasi koordinat saja
                wilayah_info = {
                    'alamat': alamat,
                    'koordinat': {'lat': lat, 'lon': lon} if lat and lon else None,
                    'status': laporan.status,
                    'id_laporan': laporan.idLaporan,
                    'nama_pelapor': laporan.nama,
                    'tanggal_lapor': laporan.tanggal_lapor.strftime('%Y-%m-%d') if laporan.tanggal_lapor else None
                }
                
                wilayah_data.append(wilayah_info)
                
            except (TypeError, ValueError, AttributeError):
                continue
        
        return wilayah_data
    
    def ranking_wilayah_bersih_kotor(self, wilayah_data):
        from collections import defaultdict
        
        # Kelompokkan data per wilayah (gunakan ekstrak dari alamat atau koordinat grid)
        wilayah_counter = defaultdict(lambda: {
            'total_laporan': 0,
            'laporan_selesai': 0,
            'laporan_pending': 0,
            'alamat_samples': []
        })
        
        for entry in wilayah_data:
            # Gunakan bagian pertama dari alamat sebagai identifikasi wilayah
            alamat = entry.get('alamat', '')
            if alamat:
                # Coba ekstrak wilayah dari alamat (contoh: ambil 3 kata pertama)
                wilayah_parts = alamat.split()[:3]
                if wilayah_parts:
                    key = ' '.join(wilayah_parts)
                else:
                    key = 'Lokasi Tidak Jelas'
            else:
                # Jika tidak ada alamat, gunakan koordinat grid
                if entry.get('koordinat'):
                    lat = entry['koordinat'].get('lat')
                    lon = entry['koordinat'].get('lon')
                    if lat and lon:
                        key = f"Koordinat: {lat:.3f},{lon:.3f}"
                    else:
                        key = 'Lokasi Tidak Diketahui'
                else:
                    key = 'Lokasi Tidak Diketahui'
            
            wilayah_counter[key]['total_laporan'] += 1
            
            if entry.get('status') == 'selesai':
                wilayah_counter[key]['laporan_selesai'] += 1
            else:
                wilayah_counter[key]['laporan_pending'] += 1
            
            # Simpan sample alamat
            if len(wilayah_counter[key]['alamat_samples']) < 3 and alamat:
                wilayah_counter[key]['alamat_samples'].append(alamat[:50])
        
        # Hitung metrics untuk ranking
        ranking_list = []
        for wilayah, data in wilayah_counter.items():
            if data['total_laporan'] > 0:
                tingkat_penyelesaian = (data['laporan_selesai'] / data['total_laporan']) * 100
                kepadatan_laporan = data['total_laporan']
                
                ranking_list.append({
                    'wilayah': wilayah,
                    'total_laporan': data['total_laporan'],
                    'laporan_selesai': data['laporan_selesai'],
                    'laporan_pending': data['laporan_pending'],
                    'tingkat_penyelesaian': round(tingkat_penyelesaian, 1),
                    'kepadatan_laporan': kepadatan_laporan,
                    'alamat_samples': data['alamat_samples'],
                    'skor_kebersihan': self.hitung_skor_kebersihan(
                        tingkat_penyelesaian, 
                        kepadatan_laporan
                    )
                })
        
        # Urutkan: terkotor (skor kebersihan terendah) sampai terbersih
        ranking_list.sort(key=lambda x: x['skor_kebersihan'])
        
        # Tambahkan peringkat
        for i, item in enumerate(ranking_list):
            item['peringkat'] = i + 1
            item['kategori'] = self.kategori_kebersihan(item['skor_kebersihan'])
        
        return {
            'ranking_terkotor': ranking_list[:5] if len(ranking_list) >= 5 else ranking_list,
            'ranking_terbersih': ranking_list[-5:][::-1] if len(ranking_list) >= 5 else ranking_list[::-1],
            'total_wilayah_teranalisis': len(ranking_list)
        }

    def hitung_skor_kebersihan(self, tingkat_penyelesaian, kepadatan_laporan):
        """
        Hitung skor kebersihan (0-100)
        Semakin tinggi skor = semakin bersih
        """
        # Faktor 1: Tingkat penyelesaian (60%)
        skor_penyelesaian = tingkat_penyelesaian * 0.6
        
        # Faktor 2: Kepadatan laporan (40%)
        # Normalisasi: asumsi maks 50 laporan/wilayah = sangat padat
        max_density = 50
        density_normalized = min(kepadatan_laporan / max_density, 1.0)
        skor_kepadatan = (1 - density_normalized) * 40
        
        total_skor = skor_penyelesaian + skor_kepadatan
        return round(total_skor, 1)
    
    def kategori_kebersihan(self, skor):
        """Kategorikan berdasarkan skor kebersihan"""
        if skor >= 80:
            return 'Sangat Bersih'
        elif skor >= 60:
            return 'Bersih'
        elif skor >= 40:
            return 'Cukup Bersih'
        elif skor >= 20:
            return 'Kotor'
        else:
            return 'Sangat Kotor'
    
    def generate_rekomendasi(self, klasifikasi, analisis_wilayah, ranking_wilayah, efektivitas, dampak_lingkungan):
        """Generate rekomendasi berdasarkan analisis dampak lingkungan"""
        rekomendasi = []
        
        # 1. Rekomendasi berdasarkan jenis sampah berbahaya
        if dampak_lingkungan.get('peringatan'):
            for peringatan in dampak_lingkungan['peringatan']:
                if peringatan['level'] == 'sangat_tinggi':
                    rekomendasi.append({
                        'prioritas': 'sangat_tinggi',
                        'kategori': 'keamanan',
                        'rekomendasi': f'PENANGANAN SEGERA: Limbah {peringatan["jenis"].upper()}',
                        'alasan': f'Ditemukan {peringatan["jumlah"]} laporan limbah berbahaya ({peringatan["persentase"]}%)',
                        'sumber_data': 'dampak_lingkungan'
                    })
                elif peringatan['level'] == 'sedang':
                    rekomendasi.append({
                        'prioritas': 'tinggi',
                        'kategori': 'lingkungan',
                        'rekomendasi': f'Program khusus untuk sampah {peringatan["jenis"]}',
                        'alasan': f'Sampah {peringatan["jenis"]} mencapai {peringatan["persentase"]}% dari total laporan',
                        'sumber_data': 'dampak_lingkungan'
                    })
        
        # 2. Rekomendasi berdasarkan lokasi berbahaya - FIX: gunakan 'lokasi' bukan 'wilayah'
        if dampak_lingkungan.get('lokasi_berbahaya'):
            lokasi_prioritas = [l for l in dampak_lingkungan['lokasi_berbahaya'] if l.get('tingkat_risiko') == 'tinggi']
            if lokasi_prioritas:
                # Gunakan key 'lokasi' yang sesuai dengan struktur data
                lokasi_nama = lokasi_prioritas[0].get('lokasi', 'Lokasi tidak diketahui')
                rekomendasi.append({
                    'prioritas': 'tinggi',
                    'kategori': 'lokasi',
                    'rekomendasi': f'Fokuskan pembersihan di {lokasi_nama}',
                    'alasan': f'Wilayah dengan konsentrasi sampah berbahaya tertinggi',
                    'sumber_data': 'dampak_lingkungan'
                })
        
        # 3. Rekomendasi efektivitas rendah
        if efektivitas.get('tingkat_penyelesaian', 0) < 50:
            rekomendasi.append({
                'prioritas': 'tinggi',
                'kategori': 'operasional',
                'rekomendasi': 'Optimalkan tim lapangan dan tingkatkan respons time',
                'alasan': f'Tingkat penyelesaian hanya {efektivitas.get("tingkat_penyelesaian", 0):.1f}%',
                'sumber_data': 'efektivitas_penanganan'
            })
        
        # 4. Rekomendasi untuk wilayah terkotor
        if ranking_wilayah.get('ranking_terkotor'):
            wilayah_terkotor = ranking_wilayah['ranking_terkotor'][0]
            rekomendasi.append({
                'prioritas': 'sedang',
                'kategori': 'wilayah',
                'rekomendasi': f'Perhatian khusus untuk {wilayah_terkotor.get("wilayah", "wilayah terkotor")}',
                'alasan': f'Wilayah dengan skor kebersihan terendah ({wilayah_terkotor.get("skor_kebersihan", 0)})',
                'sumber_data': 'ranking_wilayah'
            })
        
        # 5. Rekomendasi edukasi
        if klasifikasi.get('detail_klasifikasi'):
            plastik_data = next((item for item in klasifikasi['detail_klasifikasi'] if item.get('jenis') == 'plastik'), None)
            if plastik_data and plastik_data.get('persentase', 0) > 20:
                rekomendasi.append({
                    'prioritas': 'sedang',
                    'kategori': 'edukasi',
                    'rekomendasi': 'Kampanye pengurangan sampah plastik',
                    'alasan': f'Sampah plastik mencapai {plastik_data.get("persentase", 0):.1f}% dari total',
                    'sumber_data': 'klasifikasi_sampah'
                })
        
        # 6. Rekomendasi umum monitoring
        if klasifikasi.get('persentase_data_terstruktur', 0) < 50:
            rekomendasi.append({
                'prioritas': 'rendah',
                'kategori': 'monitoring',
                'rekomendasi': 'Tingkatkan pengisian data jenis sampah',
                'alasan': f'Hanya {klasifikasi.get("persentase_data_terstruktur", 0):.1f}% data memiliki klasifikasi terstruktur',
                'sumber_data': 'klasifikasi_sampah'
            })
        
        return rekomendasi[:6]
    
    def hitung_efektivitas_penanganan(self, queryset):
        """
        Menghitung efektivitas penanganan laporan sampah
        """
        total = queryset.count()
        selesai = queryset.filter(status='selesai').count()
        pending = total - selesai

        return {
            'total_laporan': total,
            'laporan_selesai': selesai,
            'laporan_pending': pending,
            'tingkat_penyelesaian': round((selesai / total) * 100, 2) if total > 0 else 0
        }
